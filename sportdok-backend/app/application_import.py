"""Парсер Excel-заявок формата «ПР» (лист «Регистрация»).

1 строка = 1 человек; колонки поединков и ката → несколько Registration.
Поддерживаются раскладки со столбца A (новый шаблон) и со столбца B (старые файлы).
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

SHEET_NAME = "Регистрация"
DATA_START_ROW = 8

# Раскладка A (актуальный шаблон): A=№ B=пол C=Фамилия D=Имя E=Отчество F=ДР …
LAYOUT_A = {
    "last": 3, "first": 4, "middle": 5, "gender": 2, "birth": 6, "age": 7, "weight": 8,
    "kumite": (9, 10, 11, 12), "kata": (13, 14, 15), "rank": 16, "trainer": 17,
    "club_cell": "G4",
}
# Раскладка A старая (№, Фамилия, Имя, Отчество, пол): до смены порядка колонок
LAYOUT_A_FIO_FIRST = {
    "last": 2, "first": 3, "middle": 4, "gender": 5, "birth": 6, "age": 7, "weight": 8,
    "kumite": (9, 10, 11, 12), "kata": (13, 14, 15), "rank": 16, "trainer": 17,
    "club_cell": "G4",
}
# Раскладка B (старые заявки со столбца B): B=№ C=Фамилия … F=пол …
LAYOUT_B = {
    "last": 3, "first": 4, "middle": 5, "gender": 6, "birth": 7, "age": 8, "weight": 9,
    "kumite": (10, 11, 12, 13), "kata": (14, 15, 16), "rank": 17, "trainer": 18,
    "club_cell": "H4",
}

SPECIAL_KUMITE_CATEGORY = {
    "АБС": "абсолютная категория",
    "двоеборье": "двоеборье",
}

YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")



@dataclass
class ParsedRegistration:
    discipline: str
    category_name: str
    team_number: Optional[str] = None
    source_code: str = ""


@dataclass
class ParsedAthlete:
    row: int
    last_name: str
    first_name: str
    middle_name: Optional[str]
    gender: Optional[str]
    birth_date: Optional[date]
    age_years: Optional[str]
    weight: Optional[float]
    rank: Optional[str]
    trainer_name: Optional[str]
    club_name: Optional[str]
    registrations: list[ParsedRegistration] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


@dataclass
class ParseResult:
    club_name: Optional[str]
    athletes: list[ParsedAthlete]
    file_errors: list[str] = field(default_factory=list)

    @property
    def person_count(self) -> int:
        return sum(1 for a in self.athletes if not a.errors or a.last_name or a.first_name)

    @property
    def registration_count(self) -> int:
        return sum(len(a.registrations) for a in self.athletes if not a.errors)

    @property
    def error_count(self) -> int:
        return len(self.file_errors) + sum(len(a.errors) for a in self.athletes)


def map_entry_code(code: str) -> Optional[ParsedRegistration]:
    """ОК-50 → kumite_ok/50; ОК-ката-ренгокай → kata/ОК-ката-ренгокай."""
    raw = (code or "").strip()
    if not raw:
        return None

    lower = raw.lower()
    if "ката" in lower:
        return ParsedRegistration(discipline="kata", category_name=raw, source_code=raw)

    for prefix, discipline in (("ОК-", "kumite_ok"), ("ПК-", "kumite_pk"), ("СЗ-", "kumite_sz")):
        if raw.startswith(prefix):
            rest = raw[len(prefix) :]
            if rest.startswith("команда"):
                num = rest[len("команда") :].lstrip("-").strip() or None
                return ParsedRegistration(
                    discipline=discipline,
                    category_name="командные соревнования",
                    team_number=num,
                    source_code=raw,
                )
            category = SPECIAL_KUMITE_CATEGORY.get(rest, rest)
            return ParsedRegistration(
                discipline=discipline,
                category_name=category,
                source_code=raw,
            )
    return None


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value).rstrip("0").rstrip(".") if "." in str(value) else str(value)
    return str(value).strip()


def _is_placeholder(text: str) -> bool:
    t = (text or "").strip()
    if not t:
        return True
    return set(t) <= {"_", "—", "-", " ", "."}


def _parse_gender(value: Any) -> Optional[str]:
    s = _cell_str(value).lower()
    if not s:
        return None
    if s in ("м", "m", "male", "муж", "мужской"):
        return "male"
    if s in ("ж", "f", "w", "female", "жен", "женский"):
        return "female"
    return None


def _parse_birth(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = _cell_str(value)
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_weight(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = _cell_str(value).replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _detect_layout(ws) -> dict:
    """Определяем порядок колонок по заголовкам строки 6."""
    a6 = _cell_str(ws.cell(6, 1).value)
    b6 = _cell_str(ws.cell(6, 2).value).lower()
    c6 = _cell_str(ws.cell(6, 3).value).lower()

    if b6.startswith("№") or b6 == "n":
        return LAYOUT_B
    if b6.startswith("пол") and (c6.startswith("фам") or "фам" in c6):
        return LAYOUT_A
    if b6.startswith("фам") or "фам" in b6:
        return LAYOUT_A_FIO_FIRST
    if a6.startswith("№"):
        # по умолчанию актуальный шаблон
        return LAYOUT_A

    a8 = ws.cell(8, 1).value
    b8 = ws.cell(8, 2).value
    c8 = ws.cell(8, 3).value
    if (a8 is None or a8 == "") and isinstance(b8, (int, float)) and isinstance(c8, str):
        return LAYOUT_B
    return LAYOUT_A


def _fmt_date_ru(d: Optional[date]) -> str:
    if not d:
        return ""
    return d.strftime("%d.%m.%Y")


def fill_application_template(
    template_path: Path,
    *,
    tournament_name: str,
    location: Optional[str] = None,
    admission_date: Optional[date] = None,
    team_name: Optional[str] = None,
) -> bytes:
    """Подставить данные турнира в шаблон и вернуть .xlsx как bytes."""
    # Читаем в память — на Linux путь с кириллицей иногда ломает openpyxl.open
    raw = Path(template_path).read_bytes()
    wb = load_workbook(BytesIO(raw))
    ws = wb[SHEET_NAME]
    name = (tournament_name or "").strip() or "________________"
    loc = (location or "").strip() or "____________________"
    adm = _fmt_date_ru(admission_date) or "__________"

    ws["A2"] = f"в {name}"
    ws["A5"] = f"место проведения: {loc}"
    ws["I5"] = f"дата комиссии по допуску: {adm}"

    if team_name and not _is_placeholder(team_name):
        ws["G4"] = team_name.strip()
    for col in range(7, 16):
        ws.cell(4, col).fill = YELLOW

    buf = BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def parse_application_xlsx(data: bytes, club_name_override: Optional[str] = None) -> ParseResult:
    try:
        wb = load_workbook(BytesIO(data), data_only=True)
    except Exception:
        return ParseResult(club_name=None, athletes=[], file_errors=["Не удалось прочитать Excel-файл"])

    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        return ParseResult(
            club_name=None,
            athletes=[],
            file_errors=[f"В файле нет листа «{SHEET_NAME}»"],
        )

    ws = wb[SHEET_NAME]
    layout = _detect_layout(ws)
    club_cell = layout["club_cell"]
    header_club = _cell_str(ws[club_cell].value)
    if _is_placeholder(header_club):
        header_club = ""
    club_name = (club_name_override or header_club or None)

    athletes: list[ParsedAthlete] = []
    max_row = min(ws.max_row or DATA_START_ROW, DATA_START_ROW + 500)
    empty_streak = 0

    for row in range(DATA_START_ROW, max_row + 1):
        last_name = _cell_str(ws.cell(row, layout["last"]).value)
        first_name = _cell_str(ws.cell(row, layout["first"]).value)
        middle_raw = _cell_str(ws.cell(row, layout["middle"]).value)
        middle_name = middle_raw or None

        if not last_name and not first_name:
            empty_streak += 1
            if empty_streak >= 20 and athletes:
                break
            continue
        empty_streak = 0

        errors: list[str] = []
        if not last_name:
            errors.append("нет фамилии")
        if not first_name:
            errors.append("нет имени")

        gender_raw = ws.cell(row, layout["gender"]).value
        gender = _parse_gender(gender_raw)
        if gender_raw not in (None, "") and gender is None:
            errors.append(f"неизвестный пол «{_cell_str(gender_raw)}»")

        birth_date = _parse_birth(ws.cell(row, layout["birth"]).value)
        if birth_date is None:
            errors.append("нет или неверная дата рождения")

        age_raw = _cell_str(ws.cell(row, layout["age"]).value)
        weight = _parse_weight(ws.cell(row, layout["weight"]).value)
        rank = _cell_str(ws.cell(row, layout["rank"]).value) or None
        trainer = _cell_str(ws.cell(row, layout["trainer"]).value) or None

        registrations: list[ParsedRegistration] = []
        seen_codes: set[str] = set()
        seen_reg_keys: set[tuple[str, str]] = set()

        for col in layout["kumite"] + layout["kata"]:
            raw = _cell_str(ws.cell(row, col).value)
            if not raw:
                continue
            code_key = raw.casefold()
            if code_key in seen_codes:
                errors.append(f"дубль категории «{raw}»")
                continue
            seen_codes.add(code_key)

            mapped = map_entry_code(raw)
            if mapped is None:
                errors.append(f"неизвестный код «{raw}»")
                continue

            key = (mapped.discipline, mapped.category_name)
            if key in seen_reg_keys:
                errors.append(f"дубль категории «{raw}»")
                continue
            seen_reg_keys.add(key)
            registrations.append(mapped)

        if not errors and not registrations:
            errors.append("нет выбранных дисциплин")

        athletes.append(
            ParsedAthlete(
                row=row,
                last_name=last_name,
                first_name=first_name,
                middle_name=middle_name,
                gender=gender,
                birth_date=birth_date,
                age_years=age_raw or None,
                weight=weight,
                rank=rank,
                trainer_name=trainer,
                club_name=club_name,
                registrations=registrations if not errors else [],
                errors=errors,
            )
        )

    wb.close()

    if not athletes:
        return ParseResult(
            club_name=club_name,
            athletes=[],
            file_errors=["В файле нет строк с участниками"],
        )

    return ParseResult(club_name=club_name, athletes=athletes)


def preview_dict(result: ParseResult) -> dict:
    people_ok = [a for a in result.athletes if not a.errors]
    people_err = [a for a in result.athletes if a.errors]
    return {
        "success": len(result.file_errors) == 0,
        "club_name": result.club_name,
        "rows": len(result.athletes),
        "people": len(people_ok),
        "registrations": sum(len(a.registrations) for a in people_ok),
        "error_count": result.error_count,
        "file_errors": result.file_errors,
        "errors": [
            {
                "row": a.row,
                "name": f"{a.last_name} {a.first_name}".strip(),
                "messages": a.errors,
            }
            for a in people_err
        ],
        "sample": [
            {
                "row": a.row,
                "last_name": a.last_name,
                "first_name": a.first_name,
                "middle_name": a.middle_name,
                "gender": a.gender,
                "birth_date": str(a.birth_date) if a.birth_date else None,
                "registrations": [
                    {
                        "discipline": r.discipline,
                        "category_name": r.category_name,
                        "team_number": r.team_number,
                        "source_code": r.source_code,
                    }
                    for r in a.registrations
                ],
            }
            for a in people_ok[:15]
        ],
    }
