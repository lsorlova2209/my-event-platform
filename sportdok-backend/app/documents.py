import os
import re
import zipfile
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Flowable, KeepTogether
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from app.draw import next_power_of_two, round1_pairs_by_seed, seed_position_order

# Base14 PDF fonts (Helvetica etc.) have no Cyrillic glyphs. DejaVu Sans is
# bundled under app/fonts/ (freely licensed, same font matplotlib ships) so
# PDF export works regardless of what fonts happen to be installed on the
# deployment host.
_FONT_DIR = os.path.join(os.path.dirname(__file__), "fonts")
pdfmetrics.registerFont(TTFont("DejaVuSans", os.path.join(_FONT_DIR, "DejaVuSans.ttf")))
pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", os.path.join(_FONT_DIR, "DejaVuSans-Bold.ttf")))

_NA = "#Н/Д"


def _register_times_new_roman():
    """Times New Roman для протокола сетки (эталон ФВКР); иначе DejaVuSans."""
    pairs = [
        (os.path.join(_FONT_DIR, "times.ttf"), os.path.join(_FONT_DIR, "timesbd.ttf")),
        (r"C:\Windows\Fonts\times.ttf", r"C:\Windows\Fonts\timesbd.ttf"),
        (r"C:\Windows\Fonts\timesnr.ttf", r"C:\Windows\Fonts\timesbd.ttf"),
    ]
    for reg, bold in pairs:
        if os.path.isfile(reg):
            pdfmetrics.registerFont(TTFont("TimesNewRoman", reg))
            if os.path.isfile(bold):
                pdfmetrics.registerFont(TTFont("TimesNewRoman-Bold", bold))
            return "TimesNewRoman"
    return "DejaVuSans"


BRACKET_FONT = _register_times_new_roman()
BRACKET_FONT_BOLD = "TimesNewRoman-Bold" if BRACKET_FONT == "TimesNewRoman" else "DejaVuSans-Bold"

# Командный зачёт points scale - the ТЗ names "Командный зачёт" as a
# required document but doesn't specify a scoring formula (see Приложение
# А, п.7: "прописать в собственном положении о турнире"). Provisional
# scale pending the tournament's official регламент.
TEAM_POINTS = {1: 4, 2: 3, 3: 2, 4: 1}

DISCIPLINE_LABELS = {
    "kata": "Ката",
    "kumite_ok": "Кумитэ ОК",
    "kumite_pk": "Кумитэ ПК",
    "kumite_sz": "Кумитэ СЗ",
}

# This app targets one federation-classified sport only (see CLAUDE.md) - the
# official protocol header names it explicitly, so it's a constant rather
# than a per-tournament field.
SPORT_NAME = "ВСЕСТИЛЕВОЕ КАРАТЭ"
SPORT_CODE = "0900001411Я"

GENDER_SHORT = {"male": "м", "female": "ж"}

# "Вид программы" в официальном протоколе регистрации - короткий код вида
# кумитэ + категория через пробел (напр. "ОК 70"), а не полное название
# дисциплины - сверено с реальным образцом (docs/samples/сетка ... .xlsx,
# лист "Регистрация", колонка K). Ката уже приходит с таким префиксом в
# самом category_name (см. ТЗ 4.3, "ОК-ката-годзю-рю"/"СЗ-ката-соло"),
# поэтому короткий код нужен только для кумитэ.
DISCIPLINE_SHORT = {"kumite_ok": "ОК", "kumite_pk": "ПК", "kumite_sz": "СЗ"}

GENDER_PLURAL = {"male": "Мужчины", "female": "Женщины"}


def _program_type_label(discipline, category_name):
    if discipline == "kata":
        return category_name or ""
    short = DISCIPLINE_SHORT.get(discipline, discipline)
    return f"{short} {category_name}" if category_name else short


def _discipline_group(discipline, category_name):
    """Официальная "группа дисциплин" (Ограниченный контакт/Полный
    контакт/Средства защиты) - образец группирует ката-категории по этому
    же принципу через префикс ОК-ката-.../СЗ-ката-... в category_name
    (см. docs/samples/СВОДНАЯ_СПРАВКА...xlsx), а не отдельным полем."""
    if discipline == "kumite_pk":
        return "Полный контакт"
    if discipline == "kumite_sz":
        return "Средства защиты"
    if discipline == "kata" and (category_name or "").startswith("СЗ"):
        return "Средства защиты"
    return "Ограниченный контакт"


def _dominant_age_group(participants):
    counts = {}
    for p in participants:
        ag = p.get("age_group")
        if ag:
            counts[ag] = counts.get(ag, 0) + 1
    return max(counts, key=counts.get) if counts else None


def _section_age_label(gender, age_group):
    # "Мужчины"/"Женщины" из compute_age_group уже означают взрослую
    # группу 21+ - в официальном образце это подписано как "18+ лет"
    # (см. docs/samples/СВОДНАЯ_СПРАВКА...xlsx, "Мужчины 18+ лет").
    if age_group in ("Мужчины", "Женщины"):
        return f"{age_group} 18+ лет"
    return age_group or GENDER_PLURAL.get(gender, "")


def _fmt_date(value):
    if not value:
        return ""
    if isinstance(value, (date,)):
        return value.strftime("%d.%m.%Y")
    text = str(value)
    try:
        return date.fromisoformat(text[:10]).strftime("%d.%m.%Y")
    except ValueError:
        return text


def _short_name(p):
    last = (p.get("last_name") or "").strip()
    first = (p.get("first_name") or "").strip()
    full = f"{last} {first}".strip()
    if len(full) <= 20 or not first:
        return full
    return f"{last} {first[0]}."


def _competition_level(tournament):
    return (tournament or {}).get("competition_level") or "club"


def _participant_org(p, competition_level="club"):
    if competition_level == "region":
        return (p.get("region") or p.get("club_name") or "").strip()
    return (p.get("club_name") or p.get("region") or "").strip()


def _participant_label(p, competition_level="club", short=False):
    if not p:
        return {"seed": "", "name": "", "text": ""}
    base_name = _short_name(p) if short else (p.get("full_name") or _short_name(p))
    org = _participant_org(p, competition_level)
    name = f"{base_name} ({org})" if org else base_name
    seed = str(p.get("seed") or "")
    text = f"{seed} {name}".strip() if seed else name
    return {"seed": seed, "name": name, "text": text}


def _clip_text(canv, text, max_width):
    if not text:
        return ""
    if canv.stringWidth(text, canv._fontname, canv._fontsize) <= max_width:
        return text
    ellipsis = "..."
    clipped = text
    while clipped and canv.stringWidth(clipped + ellipsis, canv._fontname, canv._fontsize) > max_width:
        clipped = clipped[:-1]
    return (clipped + ellipsis) if clipped else ellipsis


def team_standings(placements):
    """placements: [{club_name, place}, ...] across all categories.
    Returns [{club_name, points}, ...] sorted best-first."""
    points_by_club = {}
    for p in placements:
        club = p.get("club_name") or "Без клуба"
        points_by_club[club] = points_by_club.get(club, 0) + TEAM_POINTS.get(p["place"], 0)
    return sorted(
        ({"club_name": club, "points": pts} for club, pts in points_by_club.items()),
        key=lambda e: (-e["points"], e["club_name"])
    )


def _category_label(discipline, gender, category_name):
    gender_label = {"male": "муж.", "female": "жен."}.get(gender, gender or "")
    return " / ".join(part for part in (DISCIPLINE_LABELS.get(discipline, discipline), gender_label, category_name) if part)


def _header_row(ws, row, values):
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(bold=True)


REGISTRATION_PROTOCOL_HEADERS = [
    "№", "№ жребия", "Пол", "Фамилия", "Имя", "Отчество", "Дата рождения",
    "Полных лет", "Разряд, звание", "Точный вес", "Вид программы",
    "Команда", "Регион", "Тренер", "Занятое место"
]

# Official per-category bracket workbook (Регистрация + листы сетки 3…64).
_SAMPLES_DIR = Path(__file__).resolve().parents[2] / "docs" / "samples"
_REG_DATA_START = 8
_REG_DATA_END = 71
_MONTHS_RU = (
    "", "января", "февраля", "марта", "апреля", "мая", "июня",
    "июля", "августа", "сентября", "октября", "ноября", "декабря",
)


def _parse_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    text = str(value)[:10]
    try:
        d = date.fromisoformat(text)
        return datetime(d.year, d.month, d.day)
    except ValueError:
        return None


def _fmt_date_ru_long(value):
    d = _parse_date(value)
    if not d:
        return _fmt_date(value)
    return f"{d.day} {_MONTHS_RU[d.month]} {d.year}"


def _category_program_desc(cat):
    age_label = _section_age_label(cat.get("gender"), _dominant_age_group(cat.get("participants") or []))
    if cat.get("discipline") == "kata":
        return f"{age_label}  {cat.get('category_name') or ''}".strip()
    short = DISCIPLINE_SHORT.get(cat.get("discipline"), cat.get("discipline") or "")
    return f"{age_label}  {short}-весовая категория {cat.get('category_name') or ''} кг".strip()


def _category_short_label(cat):
    age = _dominant_age_group(cat.get("participants") or []) or ""
    gender_letter = {"male": "М", "female": "Ж"}.get(cat.get("gender"), "")
    age_bit = age.replace("Мужчины", "").replace("Женщины", "").strip() or "18+"
    if cat.get("discipline") == "kata":
        return f"{gender_letter} {age_bit} {cat.get('category_name') or ''}".strip()
    short = DISCIPLINE_SHORT.get(cat.get("discipline"), "")
    return f"{gender_letter} {age_bit}{short}-{cat.get('category_name') or ''}".strip()


def _safe_filename(name):
    cleaned = re.sub(r'[<>:"/\\|?*]', "_", name)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" .")
    return cleaned[:120] or "category"


def _filename_age_part(cat):
    """Возрастная группа для имени файла: «18 +», «18-20», «16-17»…"""
    age = _dominant_age_group(cat.get("participants") or []) or ""
    if age in ("Мужчины", "Женщины") or not age:
        return "18 +"
    # «Юниорки 18-20» / «Девочки 10-11» → «18-20» / «10-11»
    m = re.search(r"(\d+)\s*[-–]\s*(\d+)", age)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    m = re.search(r"(\d+)\s*\+", age)
    if m:
        return f"{m.group(1)} +"
    return "18 +"


def _filename_category_short(cat):
    """Сокращённое название категории: «ОК 70», «СЗ 72», «ОК-ката-соло»…"""
    return _program_type_label(cat.get("discipline"), cat.get("category_name")) or "категория"


def _category_excel_filename(tournament, cat):
    """
    Схема имени (как в docs/samples):
    сетка Ж_18 + ОК 70_27.09.2025.xlsx
    = сетка + Ж/М + возрастная группа + категория сокращённо + дата турнира
    """
    gender_letter = {"male": "М", "female": "Ж"}.get(cat.get("gender"), "X")
    age_part = _filename_age_part(cat)
    prog = _filename_category_short(cat)
    date_part = _fmt_date(tournament.get("event_date")) or "date"
    return _safe_filename(f"сетка {gender_letter}_{age_part} {prog}_{date_part}.xlsx")


def _category_template_path():
    if not _SAMPLES_DIR.is_dir():
        raise FileNotFoundError(f"Нет папки шаблонов: {_SAMPLES_DIR}")
    preferred = None
    for p in sorted(_SAMPLES_DIR.glob("*.xlsx")):
        if "СВОДНАЯ" in p.name.upper() or p.name.startswith("~$"):
            continue
        if preferred is None:
            preferred = p
        if "70" in p.name:
            return p
    if preferred is None:
        raise FileNotFoundError(f"Нет xlsx-шаблона категории в {_SAMPLES_DIR}")
    return preferred


def _clear_bracket_winner_inputs(wb):
    """Убирает заполненные в образце номера победителей; формулы и листья сетки не трогает.

    Ввод результата в сетке — числа в колонках C… (F12, D14, H16…);
    колонка A = № жребия на листьях, B = места 1–4 в таблице результатов.
    Формулы проверяют пустой ввод как 0: IF(F12=0,\" \",…).
    """
    for name in wb.sheetnames:
        if not name.isdigit():
            continue
        ws = wb[name]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 1, max_col=min(ws.max_column or 1, 20)):
            for cell in row:
                if cell.column < 3:
                    continue
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    continue
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    cell.value = 0


def _select_bracket_sheet(wb, participant_count):
    """Как в образце: виден лист «Регистрация» + лист сетки = числу участников."""
    target_name = str(participant_count) if participant_count >= 3 else None
    for name in wb.sheetnames:
        if name == "Регистрация":
            wb[name].sheet_state = "visible"
            continue
        if name.isdigit():
            wb[name].sheet_state = "visible" if name == target_name else "hidden"
            continue
        # Ката/командные шаблоны — скрыты, пока не понадобятся отдельно.
        wb[name].sheet_state = "hidden"

    if target_name and target_name in wb.sheetnames:
        wb.active = wb[target_name]
    else:
        wb.active = wb["Регистрация"]


def build_category_workbook(tournament, cat):
    """Один файл Excel на категорию: шаблон ФВКР с формулами + данные категории."""
    template = _category_template_path()
    # keep_vba=False; формулы VLOOKUP/DCOUNTA на листах сетки сохраняются as-is.
    wb = load_workbook(template)
    if "Регистрация" not in wb.sheetnames:
        raise ValueError("В шаблоне нет листа «Регистрация»")
    ws = wb["Регистрация"]

    event_date = tournament.get("event_date")
    date_long = _fmt_date_ru_long(event_date)
    date_short = _fmt_date(event_date)
    name = (tournament.get("name") or "").strip()
    location = (tournament.get("location") or "").strip()

    ws["A1"] = f" {name} {date_long} г." if date_long else f" {name}"
    ws["A2"] = f"Вид спорта: {SPORT_NAME} (номер-код вида спорта - {SPORT_CODE})"
    ws["A3"] = _category_program_desc(cat)
    ws["G3"] = location
    ws["L3"] = date_long or date_short
    ws["A6"] = _category_short_label(cat)
    # D6 = DCOUNTA(...) из шаблона — не перезаписываем.

    for row in range(_REG_DATA_START, _REG_DATA_END + 1):
        for col in range(1, 16):
            ws.cell(row=row, column=col).value = None

    participants = sorted(
        cat.get("participants") or [],
        key=lambda p: (p.get("seed") is None, p.get("seed") if p.get("seed") is not None else 10**9),
    )
    place_by_reg = {p["registration_id"]: p["place"] for p in cat.get("placements") or []}
    n = len(participants)

    for i, p in enumerate(participants, start=1):
        row = _REG_DATA_START + i - 1
        if row > _REG_DATA_END:
            break
        seed = p.get("seed") if p.get("seed") is not None else i
        values = [
            i,
            seed,
            GENDER_SHORT.get(p.get("gender"), p.get("gender")),
            p.get("last_name"),
            p.get("first_name"),
            p.get("middle_name"),
            _parse_date(p.get("birth_date")),
            p.get("age_years"),
            p.get("rank"),
            p.get("weight"),
            _program_type_label(p.get("discipline"), p.get("category_name")),
            p.get("club_name"),
            p.get("region"),
            p.get("trainer_name"),
            place_by_reg.get(p.get("registration_id")),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=value)

    # Пустые строки с № жребия > N — как в образце, диапазон VLOOKUP остаётся валидным.
    for i in range(n + 1, _REG_DATA_END - _REG_DATA_START + 2):
        row = _REG_DATA_START + i - 1
        if row > _REG_DATA_END:
            break
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=i)

    ws["N73"] = tournament.get("chief_judge") or ""
    ws["N74"] = tournament.get("chief_secretary") or ""

    _clear_bracket_winner_inputs(wb)
    _select_bracket_sheet(wb, n)

    return wb


def build_category_excel_zip(tournament, categories, summary=None):
    """ZIP: сводная справка + отдельный .xlsx на каждую категорию."""
    buffer = BytesIO()
    used_names = set()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        # Сводная справка — отдельный файл, как в пакете протоколов ФВКР.
        if summary is not None and categories:
            summary_wb = build_summary_workbook(tournament, summary, categories)
            summary_name = _safe_filename(
                f"СВОДНАЯ_СПРАВКА_{_fmt_date(tournament.get('event_date')) or 'date'}.xlsx"
            )
            used_names.add(summary_name.lower())
            bio = BytesIO()
            summary_wb.save(bio)
            zf.writestr(summary_name, bio.getvalue())

        for cat in categories:
            if not cat.get("participants"):
                continue
            wb = build_category_workbook(tournament, cat)
            name = _category_excel_filename(tournament, cat)
            base, ext = os.path.splitext(name)
            candidate = name
            i = 2
            while candidate.lower() in used_names:
                candidate = f"{base}_{i}{ext}"
                i += 1
            used_names.add(candidate.lower())
            bio = BytesIO()
            wb.save(bio)
            zf.writestr(candidate, bio.getvalue())
    buffer.seek(0)
    return buffer


_SUMMARY_GROUP_ORDER = ("Ограниченный контакт", "Полный контакт", "Средства защиты")


def _summary_program_label(discipline, category_name):
    """Подпись дисциплины в сводной: «ОК-70», «ПК-60», «ОК-ката-соло»."""
    if discipline == "kata":
        return category_name or ""
    short = DISCIPLINE_SHORT.get(discipline, discipline or "")
    name = (category_name or "").strip()
    return f"{short}-{name}" if name else short


def _summary_section_sort_key(cat):
    gender = cat.get("gender") or ""
    gender_rank = 0 if gender == "male" else 1 if gender == "female" else 2
    group = _discipline_group(cat.get("discipline"), cat.get("category_name"))
    try:
        group_rank = _SUMMARY_GROUP_ORDER.index(group)
    except ValueError:
        group_rank = len(_SUMMARY_GROUP_ORDER)
    age = _section_age_label(gender, _dominant_age_group(cat.get("participants") or []))
    prog = _summary_program_label(cat.get("discipline"), cat.get("category_name"))
    return (gender_rank, group_rank, age, prog)


def build_summary_workbook(tournament, summary, categories):
    """
    Отдельный файл «СВОДНАЯ_СПРАВКА_….xlsx» по образцу ФВКР:
    шапка, участники/команды, список клубов, призёры 1–4 по группам дисциплин.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"

    bold12 = Font(bold=True, size=12)
    bold11 = Font(bold=True, size=11)
    normal12 = Font(size=12)
    normal11 = Font(size=11)
    normal9 = Font(size=9)

    ws.cell(row=4, column=2, value=f'вид спорта "{SPORT_NAME.lower()}"').font = normal12
    ws.merge_cells(start_row=4, start_column=2, end_row=5, end_column=5)

    location = (tournament.get("location") or "").strip()
    date_part = _fmt_date(tournament.get("event_date"))
    loc_line = f"{location}{' ' * 85}{date_part} г." if date_part else location
    ws.cell(row=6, column=2, value=loc_line).font = normal11
    ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=5)

    club_names = sorted({
        (p.get("club_name") or "").strip()
        for cat in categories
        for p in cat.get("participants") or []
        if (p.get("club_name") or "").strip()
    })
    participant_count = summary.get("participant_count")
    if participant_count is None:
        participant_count = sum(len(cat.get("participants") or []) for cat in categories)

    ws.cell(row=8, column=1, value="Участники").font = bold11
    ws.cell(row=8, column=2, value=participant_count).font = normal11
    ws.cell(row=9, column=1, value="Команды").font = bold11
    ws.cell(row=9, column=2, value=len(club_names)).font = normal11

    ws.cell(row=11, column=1, value="Команды:").font = bold12
    row, col = 11, 2
    for name in club_names:
        ws.cell(row=row, column=col, value=name).font = normal11
        col += 1
        if col > 5:
            col = 2
            row += 1

    row = max(row + 2, 17)
    ws.cell(row=row, column=1, value="призёры").font = bold12
    for c, label in zip((2, 3, 4, 5), ("1 место", "2 место", "3 место", "4 место")):
        ws.cell(row=row, column=c, value=label).font = bold12
    row += 1
    ws.cell(row=row, column=1, value="Наименование группы дисциплин").font = normal11
    row += 1
    ws.cell(row=row, column=1, value="Возрастная категория").font = normal11
    row += 1
    ws.cell(row=row, column=1, value="спортивная дисциплина").font = normal11
    for c in (2, 3, 4, 5):
        ws.cell(row=row, column=c, value="ФИО").font = normal11
    row += 1
    for c in (2, 3, 4, 5):
        ws.cell(row=row, column=c, value="(Команда)").font = normal11
    # колонка A строк «спортивная дисциплина»/(Команда) — как в образце, объединены
    ws.merge_cells(start_row=row - 1, start_column=1, end_row=row, end_column=1)
    row += 1

    ordered = sorted(
        [c for c in categories if c.get("participants")],
        key=_summary_section_sort_key,
    )
    sections = []
    section_index = {}
    for cat in ordered:
        group = _discipline_group(cat.get("discipline"), cat.get("category_name"))
        age = _section_age_label(cat.get("gender"), _dominant_age_group(cat.get("participants") or []))
        key = (group, age)
        if key not in section_index:
            section_index[key] = len(sections)
            sections.append((key, []))
        sections[section_index[key]][1].append(cat)

    for (group_label, age_label), cats in sections:
        ws.cell(row=row, column=1, value=f"Группа дисциплин: {group_label}").font = bold12
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1
        ws.cell(row=row, column=1, value=age_label).font = bold12
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1
        for cat in cats:
            prog = _summary_program_label(cat.get("discipline"), cat.get("category_name"))
            cell = ws.cell(row=row, column=1, value=prog)
            cell.font = normal9
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=1)
            placements_by_place = {p["place"]: p for p in cat.get("placements") or []}
            for place in (1, 2, 3, 4):
                p = placements_by_place.get(place)
                if not p:
                    continue
                ws.cell(row=row, column=1 + place, value=p.get("full_name") or "").font = normal9
                ws.cell(row=row + 1, column=1 + place, value=p.get("club_name") or "").font = normal9
            row += 2

    for col_letter, width in zip("ABCDE", (17.2, 29.2, 28.8, 28.3, 29.8)):
        ws.column_dimensions[col_letter].width = width

    return wb


def build_workbook(tournament, summary, categories, team_ranking):
    """
    tournament: {name, location, event_date, registration_closes_at, status}
    summary: {participant_count, category_count, discipline_counts: {discipline: count}}
    categories: [{
        discipline, gender, category_name,
        placements: [{place, full_name, club_name, registration_id}, ...],  # may be empty
        progress: [str, ...],  # human-readable log lines
        participants: [{registration_id, seed, last_name, first_name, middle_name,
                        gender, birth_date, age_years, rank, weight, club_name,
                        region, trainer_name, discipline, category_name}, ...]
    }, ...]
    team_ranking: [{club_name, points}, ...]
    """
    wb = Workbook()
    competition_level = _competition_level(tournament)

    # ─── СВОДНАЯ СПРАВКА ────────────────────────────────────────────────
    # Формат сверен с реальным образцом (docs/samples/СВОДНАЯ_СПРАВКА...
    # xlsx): шапка турнира, число участников/команд, список команд, затем
    # призёры - таблица с местами 1-4 в отдельных колонках (ФИО в одной
    # строке, команда строкой ниже), сгруппированная по группе дисциплин
    # (Ограниченный контакт/Полный контакт/Средства защиты) и по
    # возрастной группе.
    summary_ws = wb.active
    summary_ws.title = "Сводная справка"
    summary_ws.cell(row=4, column=2, value=f'вид спорта "{SPORT_NAME.lower()}"').font = Font(bold=True, size=12)
    summary_ws.merge_cells(start_row=4, start_column=2, end_row=5, end_column=5)
    summary_ws.cell(row=6, column=2, value=f"{tournament.get('location') or ''}          {_fmt_date(tournament.get('event_date'))} г.")
    summary_ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=5)

    club_names = sorted({
        p["club_name"]
        for cat in categories
        for p in cat["participants"]
        if p.get("club_name")
    })
    summary_ws.cell(row=8, column=1, value="Участники")
    summary_ws.cell(row=8, column=2, value=summary.get("participant_count"))
    summary_ws.cell(row=9, column=1, value="Команды")
    summary_ws.cell(row=9, column=2, value=len(club_names))

    summary_ws.cell(row=11, column=1, value="Команды:").font = Font(bold=True)
    row, col = 11, 2
    for name in club_names:
        summary_ws.cell(row=row, column=col, value=name)
        col += 1
        if col > 5:
            col = 2
            row += 1

    row += 2
    summary_ws.cell(row=row, column=1, value="призёры").font = Font(bold=True)
    for col, label in zip((2, 3, 4, 5), ("1 место", "2 место", "3 место", "4 место")):
        summary_ws.cell(row=row, column=col, value=label).font = Font(bold=True)
    row += 1

    sections = []
    section_index = {}
    for cat in categories:
        key = (_discipline_group(cat["discipline"], cat["category_name"]), _section_age_label(cat["gender"], _dominant_age_group(cat["participants"])))
        if key not in section_index:
            section_index[key] = len(sections)
            sections.append((key, []))
        sections[section_index[key]][1].append(cat)

    for (group_label, age_label), cats in sections:
        summary_ws.cell(row=row, column=1, value=f"Группа дисциплин: {group_label}").font = Font(bold=True)
        summary_ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1
        summary_ws.cell(row=row, column=1, value=age_label).font = Font(bold=True)
        summary_ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1
        for cat in cats:
            summary_ws.cell(row=row, column=1, value=_program_type_label(cat["discipline"], cat["category_name"]))
            summary_ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=1)
            placements_by_place = {p["place"]: p for p in cat["placements"]}
            for place in (1, 2, 3, 4):
                p = placements_by_place.get(place)
                if p:
                    summary_ws.cell(row=row, column=1 + place, value=p["full_name"])
                    summary_ws.cell(row=row + 1, column=1 + place, value=p["club_name"])
            row += 2
    for col, width in zip("ABCDE", (26, 24, 24, 24, 24)):
        summary_ws.column_dimensions[col].width = width

    # ─── ПРОТОКОЛ РЕГИСТРАЦИИ (сетка) ──────────────────────────────────
    # Формат сверен с реальным образцом (docs/samples/сетка ....xlsx, лист
    # "Регистрация"): шапка турнира + категории, спаренная подпись
    # "вид программы/место проведения/дата" под значениями, заголовок
    # "ПРОТОКОЛ РЕГИСТРАЦИИ" вразрядку, затем полный состав участников
    # категории (не только призёры), отсортированный по номеру жребия,
    # с местом для подписи главного судьи/секретаря внизу каждого блока.
    reg_ws = wb.create_sheet("Протокол регистрации")
    row = 1
    for cat in categories:
        if not cat["participants"]:
            continue
        age_label = _section_age_label(cat["gender"], _dominant_age_group(cat["participants"]))
        if cat["discipline"] == "kata":
            program_desc = f"{age_label}  {cat['category_name']}"
        else:
            short = DISCIPLINE_SHORT.get(cat["discipline"], cat["discipline"])
            program_desc = f"{age_label}  {short}-весовая категория {cat['category_name']} кг"

        reg_ws.cell(row=row, column=1, value=f" {tournament.get('name')} {_fmt_date(tournament.get('event_date'))} г.").font = Font(bold=True, size=12)
        reg_ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
        row += 1
        reg_ws.cell(row=row, column=1, value=f"Вид спорта: {SPORT_NAME} (номер-код вида спорта - {SPORT_CODE})")
        reg_ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
        row += 1

        reg_ws.cell(row=row, column=1, value=program_desc)
        reg_ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        reg_ws.cell(row=row, column=7, value=tournament.get("location"))
        reg_ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=10)
        reg_ws.cell(row=row, column=12, value=_fmt_date(tournament.get("event_date")))
        reg_ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=14)
        row += 1
        for col, label in ((1, "вид программы"), (7, "место проведения соревнований"), (12, "Дата проведения соревнований")):
            reg_ws.cell(row=row, column=col, value=label).font = Font(italic=True, size=9)
        reg_ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        reg_ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=10)
        reg_ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=14)
        row += 1

        title_cell = reg_ws.cell(row=row, column=1, value="П Р О Т О К О Л   Р Е Г И С Т Р А Ц И И")
        title_cell.font = Font(bold=True, size=12)
        title_cell.alignment = Alignment(horizontal="center")
        reg_ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
        row += 1

        reg_headers = REGISTRATION_PROTOCOL_HEADERS.copy()
        reg_headers[11] = "Клуб"
        reg_headers[12] = "Регион"
        _header_row(reg_ws, row, reg_headers)
        for col in range(1, 16):
            reg_ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", wrap_text=True)
        row += 1

        place_by_reg = {p["registration_id"]: p["place"] for p in cat["placements"]}
        ordered = sorted(cat["participants"], key=lambda p: (p["seed"] is None, p["seed"]))
        for i, p in enumerate(ordered, start=1):
            values = [
                i, p.get("seed"), GENDER_SHORT.get(p.get("gender"), p.get("gender")),
                p.get("last_name"), p.get("first_name"), p.get("middle_name"),
                _fmt_date(p.get("birth_date")), p.get("age_years"), p.get("rank"),
                p.get("weight"), _program_type_label(p.get("discipline"), p.get("category_name")),
                p.get("club_name"), p.get("region"), p.get("trainer_name"),
                place_by_reg.get(p["registration_id"]),
            ]
            for col, value in enumerate(values, start=1):
                reg_ws.cell(row=row, column=col, value=value)
            row += 1

        row += 1
        reg_ws.cell(row=row, column=6, value="Главный судья:")
        reg_ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
        reg_ws.cell(row=row, column=8, value=tournament.get("chief_judge") or "")
        row += 1
        reg_ws.cell(row=row, column=6, value="Главный секретарь:")
        reg_ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
        reg_ws.cell(row=row, column=8, value=tournament.get("chief_secretary") or "")
        row += 3
    for col, width in zip("ABCDEFGHIJKLMNO", (4, 9, 5, 14, 14, 14, 12, 8, 12, 9, 14, 22, 16, 20, 8)):
        reg_ws.column_dimensions[col].width = width

    # ─── ПРОТОКОЛ ХОДА СОРЕВНОВАНИЯ ─────────────────────────────────────
    progress_ws = wb.create_sheet("Протокол хода соревнования")
    row = 1
    for cat in categories:
        if not cat["progress"]:
            continue
        progress_ws.cell(row=row, column=1, value=_category_label(cat["discipline"], cat["gender"], cat["category_name"])).font = Font(bold=True, size=12)
        row += 1
        for line in cat["progress"]:
            progress_ws.cell(row=row, column=1, value=line)
            row += 1
        row += 1
    progress_ws.column_dimensions["A"].width = 90

    # ─── КОМАНДНЫЙ ЗАЧЁТ ────────────────────────────────────────────────
    team_ws = wb.create_sheet("Командный зачёт")
    _header_row(team_ws, 1, ["Клуб", "Очки"])
    for i, t in enumerate(team_ranking, start=2):
        team_ws.cell(row=i, column=1, value=t["club_name"])
        team_ws.cell(row=i, column=2, value=t["points"])
    for col, width in zip("AB", (30, 10)):
        team_ws.column_dimensions[col].width = width

    return wb


def _pdf_styles():
    styles = getSampleStyleSheet()
    for name in ("Normal", "BodyText"):
        styles[name].fontName = "DejaVuSans"
    for name in ("Title", "Heading1", "Heading2", "Heading3"):
        styles[name].fontName = "DejaVuSans-Bold"
    styles["Title"].alignment = 1
    styles.add(_paragraph_style("Center", styles["Normal"], alignment=1))
    styles.add(_paragraph_style("DocTitle", styles["Normal"], fontName="DejaVuSans-Bold", fontSize=13, alignment=1))
    return styles


def _paragraph_style(name, parent, **kwargs):
    return ParagraphStyle(name, parent=parent, **kwargs)


def _make_table(rows, header=False):
    table = Table(rows, hAlign="LEFT")
    style = [
        ("FONTNAME", (0, 0), (-1, -1), "DejaVuSans"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]
    if header:
        style.append(("FONTNAME", (0, 0), (-1, 0), "DejaVuSans-Bold"))
        style.append(("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke))
    table.setStyle(TableStyle(style))
    return table


# ─── OFFICIAL PROTOCOL HEADER / FOOTER ─────────────────────────────────────
# Layout copied from the federation's own protocol forms (see docs/samples):
# tournament name, sport line, a 3-column info strip (программа / место /
# дата), the document title, then (after the body) a judge signature block.

def _spaced_caps(text):
    """«ПРОТОКОЛ ХОДА» → «П Р О Т О К О Л   Х О Д А» (как на бланке ФВКР)."""
    words = (text or "").upper().split()
    return "   ".join(" ".join(ch for ch in word) for word in words)


def _protocol_page_header(styles, tournament, program_label, doc_title, compact=False):
    """Единая шапка всех PDF-протоколов:
    название+дата · вид спорта · категория|место|дата · заголовок вразрядку.
    """
    t_name = tournament.get("name") or "Турнир"
    t_date = _fmt_date(tournament.get("event_date"))
    title = f"{t_name} {t_date} г." if t_date else t_name
    location = tournament.get("location") or ""
    date_str = t_date or ""
    program = program_label or ""

    title_fs = 11 if compact else 14
    sport_fs = 9 if compact else 12
    info_fs = 8 if compact else 11
    proto_fs = 11 if compact else 14
    gap1 = 0.15 * cm if compact else 0.35 * cm
    gap2 = 0.12 * cm if compact else 0.25 * cm
    gap3 = 0.15 * cm if compact else 0.35 * cm
    gap4 = 0.2 * cm if compact else 0.45 * cm

    title_style = ParagraphStyle(
        "ProtoMainTitle", parent=styles["Normal"],
        fontName=BRACKET_FONT_BOLD, fontSize=title_fs, alignment=1, leading=title_fs + 2,
    )
    sport_style = ParagraphStyle(
        "ProtoSport", parent=styles["Normal"],
        fontName=BRACKET_FONT, fontSize=sport_fs, alignment=1, leading=sport_fs + 2,
    )
    info_l = ParagraphStyle(
        "ProtoInfoL", parent=styles["Normal"],
        fontName=BRACKET_FONT, fontSize=info_fs, alignment=0, leading=info_fs + 2,
    )
    info_c = ParagraphStyle(
        "ProtoInfoC", parent=styles["Normal"],
        fontName=BRACKET_FONT, fontSize=info_fs, alignment=1, leading=info_fs + 2,
    )
    info_r = ParagraphStyle(
        "ProtoInfoR", parent=styles["Normal"],
        fontName=BRACKET_FONT, fontSize=info_fs, alignment=2, leading=info_fs + 2,
    )
    info_table = Table(
        [[
            Paragraph(program, info_l),
            Paragraph(location, info_c),
            Paragraph(date_str, info_r),
        ]],
        colWidths=["40%", "35%", "25%"],
    )
    info_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
    ]))
    proto_style = ParagraphStyle(
        "ProtoDocTitle", parent=styles["Normal"],
        fontName=BRACKET_FONT, fontSize=proto_fs, alignment=1, leading=proto_fs + 2,
    )
    return [
        Paragraph(title, title_style),
        Spacer(1, gap1),
        Paragraph(
            f"Вид спорта: {SPORT_NAME} (номер-код вида спорта - {SPORT_CODE})",
            sport_style,
        ),
        Spacer(1, gap2),
        info_table,
        Spacer(1, gap3),
        Paragraph(_spaced_caps(doc_title), proto_style),
        Spacer(1, gap4),
    ]


def _official_header(styles, tournament, program_label, doc_title):
    """Шапка итогового / ката-протокола — тот же вид, что у сетки кумитэ."""
    return _protocol_page_header(styles, tournament, program_label, doc_title, compact=False)


def _signature_block(styles, tournament=None):
    judge = (tournament or {}).get("chief_judge") or ""
    secretary = (tournament or {}).get("chief_secretary") or ""
    table = Table(
        [
            ["Главный судья:", "", judge],
            ["Главный секретарь:", "", secretary],
        ],
        colWidths=[4.0 * cm, 8 * cm, 8 * cm],
        hAlign="LEFT",
    )
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), BRACKET_FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("LINEBELOW", (1, 0), (1, 0), 0.5, colors.black),
        ("LINEBELOW", (1, 1), (1, 1), 0.5, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
        ("TOPPADDING", (0, 0), (-1, -1), 14),
        ("ALIGN", (2, 0), (2, -1), "LEFT"),
    ]))
    return table


# ─── ИТОГОВЫЙ ПРОТОКОЛ (full participant table) ────────────────────────────

_FULL_PROTOCOL_HEADER = [
    "№<br/>п/п", "№<br/>квал", "пол", "Фамилия", "Имя", "Отчество", "Дата<br/>рождения",
    "Полных<br/>лет", "Разряд,<br/>звание", "Вид<br/>программы", "Клуб /<br/>регион", "Тренер", "Занятое<br/>место",
]
_FULL_PROTOCOL_WIDTHS = [0.9, 1.0, 0.8, 2.8, 2.4, 2.6, 2.0, 1.3, 2.0, 2.2, 3.85, 3.85, 1.5]


def _full_protocol_table(cat, tournament):
    place_by_reg = {p["registration_id"]: p["place"] for p in cat["placements"] if p.get("registration_id")}
    header_style = ParagraphStyle("ProtoHeader", fontName="DejaVuSans-Bold", fontSize=7.5, alignment=1, leading=9)
    competition_level = _competition_level(tournament)

    def sort_key(p):
        place = place_by_reg.get(p["registration_id"])
        return (0, place) if place is not None else (1, p.get("seed") or 9999, p.get("last_name") or "")

    participants = sorted(cat["participants"], key=sort_key)
    rows = [[Paragraph(h, header_style) for h in _FULL_PROTOCOL_HEADER]]
    for i, p in enumerate(participants, start=1):
        place = place_by_reg.get(p["registration_id"])
        rows.append([
            str(i),
            str(p["seed"]) if p.get("seed") else "",
            GENDER_SHORT.get(p.get("gender"), p.get("gender") or ""),
            p.get("last_name") or "",
            p.get("first_name") or "",
            p.get("middle_name") or "",
            _fmt_date(p.get("birth_date")),
            str(p.get("age_years") or ""),
            p.get("rank") or "",
            cat["category_name"] or "",
            _participant_org(p, competition_level),
            p.get("trainer_name") or "",
            str(place) if place is not None else "",
        ])

    table = Table(rows, colWidths=[w * cm for w in _FULL_PROTOCOL_WIDTHS], repeatRows=1)
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "DejaVuSans"),
        ("FONTNAME", (0, 0), (-1, 0), "DejaVuSans-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (2, -1), "CENTER"),
        ("ALIGN", (6, 0), (7, -1), "CENTER"),
        ("ALIGN", (-1, 0), (-1, -1), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    return table


# ─── ПРОТОКОЛ ХОДА СОРЕВНОВАНИЯ — ката (round score table) ─────────────────

_KATA_ROUND_TITLES = {"round1": "1-й круг", "round2": "2-й круг", "final": "Финал"}
_KATA_ROUND_ORDER = ["round1", "round2", "final"]


def _kata_rounds_table(cat):
    """Всегда три круга (1-й / 2-й / Финал): Ката | 1–5 | Итог."""
    rounds = list(_KATA_ROUND_ORDER)
    score_by_reg_round = {(s["registration_id"], s["round_label"]): s for s in (cat.get("kata_scores") or [])}
    place_by_reg = {p["registration_id"]: p["place"] for p in cat.get("placements") or [] if p.get("registration_id")}
    competition_level = cat.get("competition_level") or "club"
    org_label = "Регион" if competition_level == "region" else "Команда"
    cols_per_round = 7  # Ката + 5 оценок + Итог

    participants = sorted(
        cat.get("participants") or [],
        key=lambda p: (
            place_by_reg.get(p["registration_id"]) is None,
            place_by_reg.get(p["registration_id"], 9999),
            p.get("seed") or 9999,
            p.get("last_name") or "",
        ),
    )
    if not participants:
        return None

    header_top = ["№", f"Фамилия Имя ({org_label})"]
    header_sub = ["", ""]
    for r in rounds:
        header_top += [_KATA_ROUND_TITLES[r]] + [""] * (cols_per_round - 1)
        header_sub += ["Ката", "1", "2", "3", "4", "5", "Итог"]
    header_top.append("Место")
    header_sub.append("")

    rows = [header_top, header_sub]
    for i, p in enumerate(participants, start=1):
        row = [str(i), _full_name_with_org(p, competition_level)]
        for r in rounds:
            s = score_by_reg_round.get((p["registration_id"], r))
            if s:
                row += [
                    s.get("kata_name") or "—",
                    f'{s["score_1"]:g}', f'{s["score_2"]:g}', f'{s["score_3"]:g}',
                    f'{s["score_4"]:g}', f'{s["score_5"]:g}', f'{s["total_score"]:g}',
                ]
            else:
                row += ["—", "", "", "", "", "", ""]
        place = place_by_reg.get(p["registration_id"])
        row.append(str(place) if place is not None else "")
        rows.append(row)

    # Три круга × 7 колонок — уместить в landscape A4
    kata_w = 1.55 * cm
    score_w = 0.82 * cm
    total_w = 0.95 * cm
    round_widths = [kata_w] + [score_w] * 5 + [total_w]
    col_widths = [0.65 * cm, 3.4 * cm] + round_widths * len(rounds) + [1.0 * cm]
    table = Table(rows, colWidths=col_widths, repeatRows=2)

    style = [
        ("FONTNAME", (0, 0), (-1, -1), "DejaVuSans"),
        ("FONTNAME", (0, 0), (-1, 1), "DejaVuSans-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 6.0),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 1), colors.whitesmoke),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (2, 0), (-1, -1), "CENTER"),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 1),
        ("RIGHTPADDING", (0, 0), (-1, -1), 1),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("SPAN", (0, 0), (0, 1)),
        ("SPAN", (1, 0), (1, 1)),
        ("SPAN", (-1, 0), (-1, 1)),
    ]
    col = 2
    for _ in rounds:
        style.append(("SPAN", (col, 0), (col + cols_per_round - 1, 0)))
        col += cols_per_round
    table.setStyle(TableStyle(style))
    return table


# ─── ПРОТОКОЛ ХОДА СОРЕВНОВАНИЙ — кумитэ (официальная сетка) ─────────────────

def _bracket_program_desc(cat):
    age_label = _section_age_label(cat["gender"], _dominant_age_group(cat["participants"]))
    if cat["discipline"] == "kata":
        return f"{age_label}  {cat['category_name']}"
    short = DISCIPLINE_SHORT.get(cat["discipline"], cat["discipline"])
    return f"{age_label}  {short}-весовая категория {cat['category_name']} кг"


def _bracket_protocol_header(styles, tournament, cat, compact=False):
    """Шапка протокола сетки — единый вид всех протоколов."""
    return _protocol_page_header(
        styles,
        tournament,
        _bracket_program_desc(cat),
        "ПРОТОКОЛ ХОДА СОРЕВНОВАНИЙ",
        compact=compact,
    )


def _full_name_with_org(p, competition_level="club"):
    if not p:
        return ""
    fio = p.get("full_name") or _short_name(p)
    org = _participant_org(p, competition_level)
    return f"{fio} ({org})" if org else fio


def _official_bracket_results_table(cat, avail_width=None, compact=False):
    """Таблица «Результаты»: Место | ФИО (команда) — компактно слева, не на всю ширину."""
    competition_level = cat.get("competition_level") or "club"
    org_label = "Регион" if competition_level == "region" else "Команда"
    by_place = {p["place"]: p for p in cat.get("placements") or []}

    rows = [
        ["Место", f"Фамилия Имя Отчество ({org_label})"],
    ]
    for place in (1, 2, 3, 4):
        p = by_place.get(place)
        if p:
            rows.append([str(place), _full_name_with_org(p, competition_level)])
        else:
            rows.append([str(place), _NA])

    # Как на бланке: таблица ~половина листа, слева; остальное место свободно
    max_table_w = 13.5 * cm
    width = min(avail_width or max_table_w, max_table_w)
    place_w = 1.6 * cm if compact else 1.8 * cm
    name_w = max(width - place_w, 8 * cm)
    font_size = 9 if compact else 11
    title_size = 10 if compact else 12
    pad = 2 if compact else 3

    table = Table(rows, colWidths=[place_w, name_w], hAlign="LEFT")
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), BRACKET_FONT),
        ("FONTNAME", (0, 0), (-1, 0), BRACKET_FONT_BOLD),
        ("FONTSIZE", (0, 0), (-1, -1), font_size),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (1, 0), (1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), pad),
        ("RIGHTPADDING", (0, 0), (-1, -1), pad),
        ("TOPPADDING", (0, 0), (-1, -1), pad),
        ("BOTTOMPADDING", (0, 0), (-1, -1), pad),
    ]))
    return KeepTogether([
        Paragraph("Результаты", ParagraphStyle(
            "ResTitle", fontName=BRACKET_FONT_BOLD, fontSize=title_size, alignment=0,
        )),
        Spacer(1, 0.1 * cm),
        table,
    ])


def _bracket_signature_block(tournament=None, avail_width=None, compact=False):
    """Подписи: слева должность, линия, справа ФИО из карточки турнира."""
    # Подписи шире таблицы результатов — место под ФИО справа
    width = min(avail_width or (22 * cm), 22 * cm)
    label_w = 3.8 * cm if compact else 4.2 * cm
    name_w = 5.5 * cm if compact else 6.5 * cm
    line_w = max(width - label_w - name_w, 4 * cm)
    font_size = 9 if compact else 11
    pad = 6 if compact else 8
    judge = (tournament or {}).get("chief_judge") or ""
    secretary = (tournament or {}).get("chief_secretary") or ""
    table = Table(
        [
            ["Главный судья:", "", judge],
            ["Главный секретарь:", "", secretary],
        ],
        colWidths=[label_w, line_w, name_w],
        hAlign="LEFT",
    )
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), BRACKET_FONT),
        ("FONTSIZE", (0, 0), (-1, -1), font_size),
        ("LINEBELOW", (1, 0), (1, 0), 0.5, colors.black),
        ("LINEBELOW", (1, 1), (1, 1), 0.5, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
        ("TOPPADDING", (0, 0), (-1, -1), pad),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("ALIGN", (2, 0), (2, -1), "LEFT"),
    ]))
    return table


def _bouts_index(bouts):
    index = {}
    for b in bouts:
        if b.get("registration_id_a") and b.get("registration_id_b"):
            index[frozenset((b["registration_id_a"], b["registration_id_b"]))] = b
    return index


def _resolve_match(a, b, bouts_by_pair):
    """a, b: participant dict or None. Returns (winner_or_None, bout_or_None).
    Matching is purely by which two registrations a bout lists, so it works
    regardless of how a round was labelled when the secretary entered it."""
    if a and not b:
        return a, None
    if b and not a:
        return b, None
    if not a and not b:
        return None, None
    bout = bouts_by_pair.get(frozenset((a["registration_id"], b["registration_id"])))
    if bout and bout["status"] == "completed" and bout.get("winner_registration_id"):
        winner = a if bout["winner_registration_id"] == a["registration_id"] else b
        return winner, bout
    return None, bout


def _bracket_rounds(participants, bouts_by_pair):
    sorted_p = sorted((p for p in participants if p.get("seed")), key=lambda p: p["seed"])
    n = len(sorted_p)
    if n == 0:
        return []
    by_local = {i + 1: p for i, p in enumerate(sorted_p)}
    size = next_power_of_two(n)
    pairs = round1_pairs_by_seed(size)
    current = []
    for la, lb in pairs:
        pa, pb = by_local.get(la), by_local.get(lb)
        winner, bout = _resolve_match(pa, pb, bouts_by_pair)
        current.append({"a": pa, "b": pb, "winner": winner, "bout": bout})
    rounds = [current]
    while len(current) > 1:
        nxt = []
        for i in range(0, len(current), 2):
            wa, wb = current[i]["winner"], current[i + 1]["winner"]
            winner, bout = _resolve_match(wa, wb, bouts_by_pair) if wa and wb else (None, None)
            nxt.append({"a": wa, "b": wb, "winner": winner, "bout": bout})
        rounds.append(nxt)
        current = nxt
    return rounds


def _loser_of(match):
    if not match or not match.get("winner"):
        return None
    w = match["winner"]
    if match.get("a") and match["a"]["registration_id"] == w["registration_id"]:
        return match.get("b")
    return match.get("a")


def _build_repechage_bracket(round1, bouts_by_pair):
    losers = [
        _loser_of(m)
        for m in round1
        if m.get("a") and m.get("b") and m.get("winner")
    ]
    losers = [p for p in losers if p]
    if not losers:
        return {"rounds": [], "champion": None}
    if len(losers) == 1:
        return {"rounds": [], "champion": losers[0]}
    reseeded = []
    for i, p in enumerate(sorted(losers, key=lambda x: x.get("seed") or 999)):
        entry = dict(p)
        entry["seed"] = i + 1
        reseeded.append(entry)
    rounds = _bracket_rounds(reseeded, bouts_by_pair)
    champion = rounds[-1][0]["winner"] if rounds else None
    return {"rounds": rounds, "champion": champion}


def _compute_kumite_bracket_data(participants, bouts):
    """Та же логика, что computeKumiteBracketData в App.jsx."""
    reg_ids = {p["registration_id"] for p in participants}
    table_bouts = [
        b for b in bouts
        if b.get("registration_id_a") in reg_ids and b.get("registration_id_b") in reg_ids
    ]
    bouts_by_pair = _bouts_index(table_bouts)

    if len(participants) == 3:
        sorted_p = sorted(participants, key=lambda p: p.get("seed") or 99)
        pairs = []
        for i, j in ((0, 1), (0, 2), (1, 2)):
            a, b = sorted_p[i], sorted_p[j]
            winner, bout = _resolve_match(a, b, bouts_by_pair)
            pairs.append({"a": a, "b": b, "winner": winner, "bout": bout})
        return {"round_robin": True, "pairs": pairs, "drawn": any(p.get("seed") for p in participants), "participant_count": 3}

    if not any(p.get("seed") for p in participants):
        return {"round_robin": False, "drawn": False}

    by_subgroup = {}
    use_parity = len(participants) >= 5
    for p in participants:
        if use_parity:
            key = 1 if p["seed"] % 2 == 1 else 2
        else:
            key = p.get("subgroup") or 1
        by_subgroup.setdefault(key, []).append(p)

    subgroup_keys = sorted(by_subgroup.keys())
    rounds_per_group = [_bracket_rounds(by_subgroup[k], bouts_by_pair) for k in subgroup_keys]
    rounds_per_group = [r for r in rounds_per_group if r]
    two_groups = len(rounds_per_group) > 1

    champs = [r[-1][0]["winner"] if r else None for r in rounds_per_group]
    final_match = None
    if two_groups:
        final_match = {"a": champs[0], "b": champs[-1], "winner": None}
        if champs[0] and champs[-1]:
            final_match["winner"], _ = _resolve_match(champs[0], champs[-1], bouts_by_pair)

    repechage_per_group = []
    bronze_match = None
    if two_groups:
        repechage_per_group = [
            _build_repechage_bracket(r[0] if r else [], bouts_by_pair)
            for r in rounds_per_group
        ]
        rep_champs = [g["champion"] for g in repechage_per_group]
        bronze_match = {"a": rep_champs[0], "b": rep_champs[-1], "winner": None}
        if rep_champs[0] and rep_champs[-1]:
            bronze_match["winner"], _ = _resolve_match(
                rep_champs[0], rep_champs[-1], bouts_by_pair)
    elif rounds_per_group and len(rounds_per_group[0]) >= 2:
        semi = rounds_per_group[0][-2]
        la = _loser_of(semi[0]) if len(semi) >= 1 else None
        lb = _loser_of(semi[1]) if len(semi) >= 2 else None
        bronze_match = {"a": la, "b": lb, "winner": None}
        if la and lb:
            bronze_match["winner"], _ = _resolve_match(la, lb, bouts_by_pair)

    return {
        "round_robin": False,
        "drawn": True,
        "rounds_per_group": rounds_per_group,
        "two_groups": two_groups,
        "final_match": final_match,
        "bronze_match": bronze_match,
        "repechage_per_group": repechage_per_group,
        "participant_count": len(participants),
    }


def _excel_bracket_style(bracket_data):
    """Одинаковая логика размеров для всех сеток; 17+ — зеркало, иначе слева→направо."""
    groups = bracket_data.get("rounds_per_group") or [[[]]]
    if bracket_data.get("round_robin"):
        row_leaves = 2
        two_groups = False
        n_part = 3
        total_leaves = 2
    else:
        per_group = [2 * len(g[0]) for g in groups if g]
        total_leaves = sum(per_group) if per_group else 0
        two_groups = bool(bracket_data.get("two_groups"))
        n_part = bracket_data.get("participant_count")
        if n_part is None:
            n_part = sum(
                1
                for g in groups if g
                for m in g[0]
                for p in (m.get("a"), m.get("b"))
                if p
            )
        # Плотность по числу рядов по вертикали:
        # зеркало — максимум в одной половине; до 16 — обе подгруппы друг под другом.
        if two_groups and n_part >= 17:
            row_leaves = max(per_group or [0])
        elif two_groups:
            row_leaves = total_leaves
        else:
            row_leaves = max(per_group or [0])

    # База одна и та же для 3…18+ (дальше единый fit под страницу, как у 2–16).
    layout = (
        "round_robin" if bracket_data.get("round_robin")
        else "horizontal" if (two_groups and n_part >= 17)
        else "single"
    )
    box_w = 170.0
    seed_w = 14.0
    # Зеркало: плотнее по горизонтали — ячейки крупнее после fit под страницу
    h_gap = 8.0 if layout == "horizontal" else 24.0
    if row_leaves > 12:
        box_h = 12.5 if layout == "horizontal" else 11.0
        row_gap = 6.5 if layout == "horizontal" else 7.5
        pair_gap = 4.5 if layout == "horizontal" else 5.5
        seed_font, name_font, label_font = (6.5, 8.5, 8.5) if layout == "horizontal" else (5.5, 7.5, 7.5)
    elif row_leaves > 8:
        box_h = 12.5
        row_gap = 7.5
        pair_gap = 7.5
        seed_font, name_font, label_font = 7.0, 9.0, 9.0
    else:
        box_h = 14.5
        row_gap = 8.0
        pair_gap = 10.0
        seed_font, name_font, label_font = 8.0, 10.0, 10.0
    row_h = box_h + row_gap

    return {
        "seed_w": seed_w,
        "leaf_w": box_w,
        "h_gap": h_gap,
        "round_w": box_w,
        "row_h": row_h,
        "box_h": box_h,
        "row_gap": row_gap,
        "pair_gap": pair_gap,
        "seed_font": seed_font,
        "name_font": name_font,
        "winner_font": name_font,
        "label_font": label_font,
        # Зеркало: крошечный зазор между финалистами (бокс «1-е место» сверху, не в зазоре)
        "center_gap": 6.0 if layout == "horizontal" else (h_gap + box_w + h_gap),
        "group_gap": pair_gap if (layout == "single" and two_groups) else (row_h + pair_gap),
        "layout": layout,
    }


def _scale_bracket_style(style, factor):
    """Единый коэффициент на ячейки, зазоры и шрифты (пропорции как у сеток 2–16)."""
    if abs(factor - 1.0) < 0.001:
        return style
    factor = max(min(factor, 1.75), 0.28)
    out = dict(style)
    for key in (
        "seed_w", "leaf_w", "h_gap", "round_w", "box_h", "row_gap", "pair_gap",
        "seed_font", "name_font", "winner_font", "label_font",
    ):
        if key in out:
            out[key] = out[key] * factor
    out["seed_font"] = max(out["seed_font"], 4.0)
    out["name_font"] = max(out["name_font"], 4.5)
    out["winner_font"] = max(out["winner_font"], 4.5)
    out["label_font"] = max(out["label_font"], 4.5)
    out["row_h"] = out["box_h"] + out["row_gap"]
    out["round_w"] = out["leaf_w"]
    if out.get("layout") == "horizontal":
        # Не раздувать зазор вместе с ячейками — между финалистами всегда узко
        out["center_gap"] = 6.0
    else:
        out["center_gap"] = out["h_gap"] + out["leaf_w"] + out["h_gap"]
    if style.get("layout") == "single" and style.get("group_gap", 99) <= style.get("pair_gap", 0) + 0.1:
        out["group_gap"] = out["pair_gap"]
    else:
        out["group_gap"] = out["row_h"] + out["pair_gap"]
    return out


def _bracket_col_x(c, box_w, h_gap, round_w=None):
    """Левый край колонки раунда c (0 — листья). Все колонки одной ширины box_w."""
    return c * (box_w + h_gap)


def _fit_name_font(canv, text, font_name, max_size, min_size, max_width):
    """Уменьшить шрифт под ширину поля; размер бокса не меняется."""
    size = max_size
    while size > min_size and canv.stringWidth(text or "", font_name, size) > max_width:
        size -= 0.5
    return size


def _layout_single_tree(rounds, y_start, style, leaf_bye_seeds=None, compact_finish=False):
    """Чертёж одной подгруппы: 1) боксы 1-го круга 2) соединители 3) бокс победителя.
    compact_finish (зеркало 17+): финалист в колонке полуфиналистов, ровно между
    ними по Y; вертикаль вилки справа коротких штрихом. После зеркала — справа.
    """
    box_w = style["leaf_w"]
    box_h = style["box_h"]
    row_h = style["row_h"]
    h_gap = style["h_gap"]
    pair_gap = style.get("pair_gap", 0)
    boxes, lines = [], []
    leaf_n = 2 * len(rounds[0])
    y = y_start
    ys0, present0 = [], []

    # Этап 1 — все прямоугольники первого раунда (с зазором между ячейками)
    for i in range(leaf_n):
        cy = y + box_h / 2
        ys0.append(cy)
        match = rounds[0][i // 2]
        slot = match["a"] if i % 2 == 0 else match["b"]
        present0.append(bool(slot))
        if slot:
            boxes.append({"kind": "leaf", "x": 0, "y": y, "width": box_w, "participant": slot})
        elif leaf_bye_seeds and i < len(leaf_bye_seeds):
            boxes.append({"kind": "bye", "x": 0, "y": y, "width": box_w, "seed": leaf_bye_seeds[i]})
        y += row_h
        if pair_gap and i % 2 == 1 and i < leaf_n - 1:
            y += pair_gap

    col_ys, col_present = ys0, present0
    max_rounds = len(rounds)
    fork_exit_x = 0.0
    for c in range(1, max_rounds + 1):
        next_ys = []
        x_prev_right = _bracket_col_x(c - 1, box_w, h_gap) + box_w
        is_finalist_round = compact_finish and c == max_rounds

        if is_finalist_round:
            # Финалист в той же колонке, что полуфиналисты — ровно между ними по вертикали.
            # Вертикаль вилки справа с коротким штрихом (не впритык к боксу).
            x_sf = _bracket_col_x(c - 1, box_w, h_gap)
            x_box = x_sf
            fin_right = x_box + box_w
            stub = max(h_gap * 0.2, 3)
            mid_x = fin_right + stub
            fork_exit_x = mid_x
            for j, match in enumerate(rounds[c - 1]):
                ya = col_ys[2 * j]
                yb = col_ys[2 * j + 1]
                py = (ya + yb) / 2
                next_ys.append(py)
                lines.extend([
                    (fin_right, ya, mid_x, ya),
                    (fin_right, yb, mid_x, yb),
                    (mid_x, ya, mid_x, yb),
                    (fin_right, py, mid_x, py),
                ])
                boxes.append({
                    "kind": "participant_box", "round": c,
                    "x": x_box, "y": py - box_h / 2, "width": box_w,
                    "participant": match.get("winner"),
                })
        else:
            x_next_left = _bracket_col_x(c, box_w, h_gap)
            for j, match in enumerate(rounds[c - 1]):
                ya = col_ys[2 * j]
                yb = col_ys[2 * j + 1]
                py = (ya + yb) / 2
                next_ys.append(py)
                mid_x = x_prev_right + (x_next_left - x_prev_right) / 2
                lines.extend([
                    (x_prev_right, ya, mid_x, ya),
                    (x_prev_right, yb, mid_x, yb),
                    (mid_x, ya, mid_x, yb),
                    (mid_x, py, x_next_left, py),
                ])
                boxes.append({
                    "kind": "participant_box", "round": c,
                    "x": x_next_left, "y": py - box_h / 2, "width": box_w,
                    "participant": match.get("winner"),
                })
            fork_exit_x = x_next_left + box_w
        col_ys, col_present = next_ys, [True] * len(next_ys)

    champ_y = col_ys[0] if col_ys else ys0[0]
    if compact_finish and max_rounds >= 1:
        # Короткий хвост после вертикали (без пустой «колонки» шириной с ячейку)
        finish_inset = max(h_gap * 0.15, 2)
        champ_x = fork_exit_x
        tree_w = champ_x + finish_inset
    else:
        tree_w = _bracket_col_x(max_rounds, box_w, h_gap) + box_w
        champ_x = tree_w
    return boxes, lines, y, champ_x, champ_y, tree_w


def _mirror_tree(boxes, lines, tree_w, leaf_w, origin_x):
    m_boxes = []
    for b in boxes:
        nb = dict(b)
        w = b.get("width", leaf_w)
        nb["x"] = origin_x + tree_w - b["x"] - w
        if b["kind"] in ("winner", "leaf", "participant_box", "bye"):
            nb["mirrored"] = True
        if b["kind"] == "winner":
            nb["line_from"] = origin_x + tree_w - b["line_to"]
            nb["line_to"] = origin_x + tree_w - b["line_from"]
        m_boxes.append(nb)
    m_lines = [
        (origin_x + tree_w - x2, y2, origin_x + tree_w - x1, y1)
        for x1, y1, x2, y2 in lines
    ]
    return m_boxes, m_lines


def _leaf_bye_seeds_for_rounds(rounds):
    n = sum(1 for m in rounds[0] for p in (m.get("a"), m.get("b")) if p)
    size = next_power_of_two(max(n, 1))
    return seed_position_order(size)


def _append_bronze_block(all_boxes, all_lines, labels, bronze_match, repechage, style, y):
    """Блок «поединок за 3-е место» под основной сеткой."""
    leaf_w = style["leaf_w"]
    box_h = style["box_h"]
    row_h = style["row_h"]
    h_gap = style["h_gap"]
    if not bronze_match:
        return y
    has_rep = any(g.get("rounds") for g in (repechage or []))
    labels.append({"x": 0, "y": y + row_h * 0.4, "text": "поединок за 3-е место", "bold": True})
    y += row_h * 1.6
    if has_rep:
        rep_champs = []
        for group in repechage:
            rounds = group.get("rounds") or []
            if rounds:
                rb, rl, y_end, rcx, rcy, _ = _layout_single_tree(rounds, y, style)
                all_boxes.extend(rb)
                all_lines.extend(rl)
                rep_champs.append((rcx, rcy))
                y = y_end
            elif group.get("champion"):
                py = y + box_h / 2
                all_boxes.append({
                    "kind": "participant_box", "round": 2,
                    "x": 0, "y": y, "width": leaf_w, "participant": group["champion"],
                })
                rep_champs.append((leaf_w, py))
                y += row_h
        if len(rep_champs) >= 2:
            ya, yb = rep_champs[0][1], rep_champs[-1][1]
            py = (ya + yb) / 2
            x_prev_right = leaf_w
            x_next_left = leaf_w + h_gap
            mid_x = x_prev_right + (x_next_left - x_prev_right) / 2
            all_lines.extend([
                (x_prev_right, ya, mid_x, ya), (x_prev_right, yb, mid_x, yb),
                (mid_x, ya, mid_x, yb), (mid_x, py, x_next_left, py),
            ])
            all_boxes.append({
                "kind": "participant_box", "round": 2,
                "x": x_next_left, "y": py - box_h / 2, "width": leaf_w,
                "participant": bronze_match.get("winner"),
            })
    else:
        ya = y + box_h / 2
        yb = ya + row_h
        all_boxes.append({
            "kind": "leaf", "x": 0, "y": y, "width": leaf_w,
            "participant": bronze_match.get("a"),
        })
        all_boxes.append({
            "kind": "leaf", "x": 0, "y": y + row_h, "width": leaf_w,
            "participant": bronze_match.get("b"),
        })
        py = (ya + yb) / 2
        x_prev_right = leaf_w
        x_next_left = leaf_w + h_gap
        mid_x = x_prev_right + (x_next_left - x_prev_right) / 2
        all_lines.extend([
            (x_prev_right, ya, mid_x, ya), (x_prev_right, yb, mid_x, yb),
            (mid_x, ya, mid_x, yb), (mid_x, py, x_next_left, py),
        ])
        all_boxes.append({
            "kind": "participant_box", "round": 2,
            "x": x_next_left, "y": py - box_h / 2, "width": leaf_w,
            "participant": bronze_match.get("winner"),
        })
        y = yb + box_h / 2
    return y + box_h + 8


def _shift_bracket_geometry(boxes, lines, labels, dy):
    """Сдвиг всей геометрии вниз (нужен, когда «1-е место» рисуется выше сеток)."""
    if abs(dy) < 0.01:
        return
    for b in boxes:
        b["y"] = b["y"] + dy
    for i, (x1, y1, x2, y2) in enumerate(lines):
        lines[i] = (x1, y1 + dy, x2, y2 + dy)
    for lbl in labels:
        lbl["y"] = lbl["y"] + dy


def _append_final(all_boxes, all_lines, champ_points, final_match, style, layout_w, tree_w, right_origin=None, labels=None):
    """Финал: single — бокс справа; horizontal — схождение по центру, «1-е место» сверху.
    Возвращает (layout_w, dy) — dy > 0 если сетку сдвинули вниз под бокс чемпиона.
    """
    if not final_match or len(champ_points) < 2:
        return layout_w, 0
    leaf_w = style["leaf_w"]
    box_h = style["box_h"]
    h_gap = style["h_gap"]
    center_gap = style["center_gap"]
    label_fs = max(style.get("label_font", 7), 5)

    y0, y1 = champ_points[0][1], champ_points[1][1]
    py = (y0 + y1) / 2
    if right_origin is not None:
        # «1-е место» сверху по центру — свои размеры, без линий к сетке.
        mid_x = tree_w + center_gap / 2
        label_fs = max(label_fs * 1.15, label_fs + 1)
        label_band = label_fs + 4
        # Крупнее обычных ячеек; ширина не ограничена center_gap (может заходить в пустоту сверху)
        champ_h = max(box_h * 2.1, 22.0)
        champ_w = max(leaf_w * 1.45, 120.0)
        champ_top = 0.0
        champ_bottom = champ_h
        # Не перекрывать зону финалистов по вертикали
        if champ_bottom + label_band > min(y0, y1) - 6:
            champ_bottom = min(y0, y1) - label_band - 6
            champ_top = champ_bottom - champ_h

        x_box = mid_x - champ_w / 2
        all_boxes.append({
            "kind": "participant_box", "round": 99,
            "x": x_box, "y": champ_top, "width": champ_w, "height": champ_h,
            "participant": final_match.get("winner"),
            "name_font_scale": 1.25,
        })
        if labels is not None:
            labels.append({
                "x": mid_x,
                "y": champ_bottom + label_fs * 0.9,
                "text": "1-е место",
                "bold": True,
                "align": "center",
                "font_scale": 1.2,
            })

        top_pad = max(0, -champ_top + 2)
        if top_pad:
            _shift_bracket_geometry(all_boxes, all_lines, labels or [], top_pad)
        return max(layout_w, x_box + champ_w, champ_points[1][0]), top_pad

    x_box = max(p[0] for p in champ_points) + h_gap
    mid_x = x_box - h_gap / 2
    all_lines.extend([
        (champ_points[0][0], y0, mid_x, y0),
        (champ_points[1][0], y1, mid_x, y1),
        (mid_x, y0, mid_x, y1),
        (mid_x, py, x_box, py),
    ])
    all_boxes.append({
        "kind": "participant_box", "round": 99,
        "x": x_box, "y": py - box_h / 2, "width": leaf_w,
        "participant": final_match.get("winner"),
    })
    return max(layout_w, x_box + leaf_w), 0


def _compute_bracket_layout_round_robin(bracket_data, style):
    """Круговая система (3 участника): три матча тем же чертежом, что олимпийская сетка."""
    pairs = bracket_data.get("pairs") or []
    all_boxes, all_lines, labels = [], [], []
    y = 0
    layout_w = 0
    group_gap = style.get("group_gap", style["row_h"] * 1.2)
    for i, match in enumerate(pairs):
        rounds = [[match]]
        b, l, y_end, _cx, _cy, tw = _layout_single_tree(rounds, y, style)
        all_boxes.extend(b)
        all_lines.extend(l)
        layout_w = max(layout_w, tw)
        y = y_end
        if i < len(pairs) - 1:
            y += group_gap
    return all_boxes, all_lines, labels, layout_w, y


def _compute_bracket_layout_horizontal(bracket_data, style):
    """Две подгруппы зеркально слева/справа + финал по центру."""
    leaf_w = style["leaf_w"]
    center_gap = style["center_gap"]
    rounds_per_group = bracket_data["rounds_per_group"]
    final_match = bracket_data.get("final_match")
    bronze_match = bracket_data.get("bronze_match")
    repechage = bracket_data.get("repechage_per_group") or []

    all_boxes, all_lines, labels = [], [], []
    b0, l0, y0_end, cx0, cy0, tw0 = _layout_single_tree(
        rounds_per_group[0], 0, style, _leaf_bye_seeds_for_rounds(rounds_per_group[0]),
        compact_finish=True)
    b1, l1, y1_end, cx1, cy1, tw1 = _layout_single_tree(
        rounds_per_group[1], 0, style, _leaf_bye_seeds_for_rounds(rounds_per_group[1]),
        compact_finish=True)
    tree_w = max(tw0, tw1)
    right_origin = tree_w + center_gap
    b1m, l1m = _mirror_tree(b1, l1, tw1, leaf_w, right_origin)
    all_boxes.extend(b0)
    all_boxes.extend(b1m)
    all_lines.extend(l0)
    all_lines.extend(l1m)
    y = max(y0_end, y1_end)
    # Точки выхода — правый край левого финалиста / левый край правого
    champ_points = [(cx0, cy0), (right_origin + tw1 - cx1, cy1)]
    layout_w = right_origin + tree_w
    layout_w, dy = _append_final(
        all_boxes, all_lines, champ_points, final_match, style, layout_w, tree_w, right_origin, labels)
    y = y + dy
    y = _append_bronze_block(all_boxes, all_lines, labels, bronze_match, repechage, style, y)
    return all_boxes, all_lines, labels, layout_w, y


def _compute_bracket_layout_single(bracket_data, style):
    """Слева→направо без зеркала: одна подгруппа или две друг под другом (до 16)."""
    rounds_per_group = bracket_data["rounds_per_group"]
    final_match = bracket_data.get("final_match")
    bronze_match = bracket_data.get("bronze_match")
    repechage = bracket_data.get("repechage_per_group") or []
    box_w = style["leaf_w"]
    box_h = style["box_h"]
    h_gap = style["h_gap"]
    group_gap = style.get("group_gap", style["row_h"])
    all_boxes, all_lines, labels = [], [], []
    y = 0
    champ_points = []
    layout_w = 0
    max_rounds = max((len(r) for r in rounds_per_group), default=0)

    for gi, rounds in enumerate(rounds_per_group):
        b, l, y_end, cx, cy, tw = _layout_single_tree(
            rounds, y, style, _leaf_bye_seeds_for_rounds(rounds))
        all_boxes.extend(b)
        all_lines.extend(l)
        layout_w = max(layout_w, tw)
        champ_x, champ_y = cx, cy
        # Выровнять более короткую подгруппу по числу кругов (как на экране)
        champion = rounds[-1][0].get("winner") if rounds else None
        for c in range(len(rounds) + 1, max_rounds + 1):
            x_prev_right = _bracket_col_x(c - 1, box_w, h_gap) + box_w
            x_next_left = _bracket_col_x(c, box_w, h_gap)
            all_lines.append((x_prev_right, champ_y, x_next_left, champ_y))
            all_boxes.append({
                "kind": "participant_box", "round": c,
                "x": x_next_left, "y": champ_y - box_h / 2, "width": box_w,
                "participant": champion,
            })
            champ_x = x_next_left + box_w
            layout_w = max(layout_w, champ_x)
        champ_points.append((champ_x, champ_y))
        y = y_end
        if gi < len(rounds_per_group) - 1:
            y += group_gap

    if final_match and len(champ_points) >= 2:
        layout_w, _dy = _append_final(
            all_boxes, all_lines, champ_points, final_match, style, layout_w, layout_w, None)
    y = _append_bronze_block(all_boxes, all_lines, labels, bronze_match, repechage, style, y)
    return all_boxes, all_lines, labels, layout_w, y


def _compute_bracket_layout(bracket_data, style):
    layout = style.get("layout", "single")
    if layout == "round_robin":
        return _compute_bracket_layout_round_robin(bracket_data, style)
    if layout == "horizontal":
        return _compute_bracket_layout_horizontal(bracket_data, style)
    return _compute_bracket_layout_single(bracket_data, style)


class _OfficialBracketDiagram(Flowable):
    """Официальный бланк сетки: сначала линии, потом прямоугольники, потом текст.

    Ячейки/шрифты/отступы логически уменьшаются, чтобы сетка влезла в avail_*.
    """

    def __init__(self, bracket_data, avail_width, avail_height, competition_level="club", font_name=None):
        super().__init__()
        self.competition_level = competition_level
        self.font_name = font_name or BRACKET_FONT
        self.font_bold = BRACKET_FONT_BOLD
        max_w = max((avail_width or 0) * 0.98, 1.0)
        max_h = max(avail_height or 0, 1.0)

        style = _excel_bracket_style(bracket_data)
        boxes, lines, labels, lw, lh = _compute_bracket_layout(bracket_data, style)

        # Та же логика, что у сеток 2–16:
        # 1) подгонка по высоте (ячейки/шрифты/зазоры пропорционально)
        if lh > max_h and lh > 0:
            style = _scale_bracket_style(style, max_h / lh)
            boxes, lines, labels, lw, lh = _compute_bracket_layout(bracket_data, style)

        # 2) если шире страницы — жмём ширину ячейки и горизонтальные зазоры
        if lw > max_w and lw > 0:
            hx = max_w / lw
            style = dict(style)
            for key in ("seed_w", "leaf_w", "h_gap", "round_w"):
                style[key] = style[key] * hx
            style["round_w"] = style["leaf_w"]
            if style.get("layout") == "horizontal":
                style["center_gap"] = 6.0
            else:
                style["center_gap"] = style["h_gap"] + style["leaf_w"] + style["h_gap"]
            boxes, lines, labels, lw, lh = _compute_bracket_layout(bracket_data, style)

        # 3) Добор свободного места под лист (как у небольших сеток 2–16)
        if lw > 0 and lh > 0:
            grow = min(max_w / lw, max_h / lh)
            if grow > 1.02:
                style = _scale_bracket_style(style, grow)
                boxes, lines, labels, lw, lh = _compute_bracket_layout(bracket_data, style)
        # Ширина упирается в край, по высоте есть место — растим только ряды/кегль
        if lw > 0 and lh > 0 and lh < max_h * 0.97 and lw >= max_w * 0.94:
            vy = min(max_h / lh, 1.75)
            compact_groups = (
                style.get("layout") == "single"
                and style.get("group_gap", 99) <= style.get("pair_gap", 0) + 0.1
            )
            style = dict(style)
            for key in (
                "box_h", "row_gap", "pair_gap",
                "seed_font", "name_font", "winner_font", "label_font",
            ):
                style[key] = style[key] * vy
            style["seed_font"] = max(style["seed_font"], 4.0)
            style["name_font"] = max(style["name_font"], 4.5)
            style["winner_font"] = max(style["winner_font"], 4.5)
            style["label_font"] = max(style["label_font"], 4.5)
            style["row_h"] = style["box_h"] + style["row_gap"]
            style["group_gap"] = style["pair_gap"] if compact_groups else (
                style["row_h"] + style["pair_gap"]
            )
            boxes, lines, labels, lw, lh = _compute_bracket_layout(bracket_data, style)

        # 4) финальный canvas-scale, если чуть вылезли
        s = 1.0
        if lw > max_w or lh > max_h:
            s = min(max_w / lw if lw else 1.0, max_h / lh if lh else 1.0)

        self._style = style
        self._boxes, self._lines, self._labels = boxes, lines, labels
        self.scale_x = s
        self.scale_y = s
        self.scale = s
        self.width = lw * s
        self.height = lh * s
        self.box_w = style["leaf_w"] * s
        self.box_h = style["box_h"] * s
        self.seed_w = style["seed_w"] * s
        self._layout_box_h = style["box_h"]
        font_scale = max(s, 0.45)
        self._seed_font = max(4.0, style["seed_font"] * font_scale)
        self._name_font = max(4.5, style["name_font"] * font_scale)
        self._label_font = max(4.5, style["label_font"] * font_scale)

    def wrap(self, avail_width, avail_height):
        return self.width, self.height

    def _ly(self, y_layout):
        return self.height - y_layout * self.scale_y

    def _box_rect(self, item):
        """Экранные координаты прямоугольника (ширина/высота из layout)."""
        layout_w = item.get("width", self._style["leaf_w"])
        layout_h = item.get("height", self._style["box_h"])
        w = layout_w * self.scale_x
        h = layout_h * self.scale_y
        x = item["x"] * self.scale_x
        y_c = self._ly(item["y"] + layout_h / 2)
        return x, y_c - h / 2, w, h, y_c

    def _draw_box_outline(self, item):
        """Бокс № | ФИО — как на эталоне после соединителя."""
        x, y_b, w, h, _ = self._box_rect(item)
        canv = self.canv
        canv.setLineWidth(0.5)
        canv.rect(x, y_b, w, h, stroke=1, fill=0)
        sw = min(self.seed_w, w * 0.18)
        if item.get("mirrored"):
            canv.line(x + w - sw, y_b, x + w - sw, y_b + h)
        else:
            canv.line(x + sw, y_b, x + sw, y_b + h)

    def _draw_box_text(self, item):
        x, y_b, w, h, y_c = self._box_rect(item)
        canv = self.canv
        sw = min(self.seed_w, w * 0.18)
        name_max_w = w - sw - 4
        mirrored = item.get("mirrored", False)
        if item["kind"] == "bye":
            seed_text = str(item.get("seed") or "")
            name_text = ""
        elif item.get("participant"):
            label = _participant_label(item["participant"], self.competition_level, short=True)
            seed_text, name_text = label["seed"], label["name"]
        else:
            seed_text, name_text = "", _NA
        base_name = self._name_font * float(item.get("name_font_scale") or 1.0)
        name_fs = _fit_name_font(
            canv, name_text, self.font_name, base_name, 5.0, name_max_w) if name_text else base_name
        seed_fs = min(self._seed_font * float(item.get("name_font_scale") or 1.0), name_fs)
        ty = y_c - name_fs / 3
        if mirrored:
            if seed_text:
                canv.setFont(self.font_name, seed_fs)
                canv.drawCentredString(x + w - sw / 2, ty, seed_text)
            if name_text:
                canv.setFont(self.font_name, name_fs)
                canv.drawRightString(x + w - sw - 2, ty, _clip_text(canv, name_text, name_max_w))
        else:
            if seed_text:
                canv.setFont(self.font_name, seed_fs)
                canv.drawCentredString(x + sw / 2, ty, seed_text)
            if name_text:
                canv.setFont(self.font_name, name_fs)
                canv.drawString(x + sw + 2, ty, _clip_text(canv, name_text, name_max_w))

    def draw(self):
        canv = self.canv
        # 1) Все соединители
        canv.setLineWidth(0.5)
        for x1, y1, x2, y2 in self._lines:
            canv.line(
                x1 * self.scale_x, self._ly(y1),
                x2 * self.scale_x, self._ly(y2),
            )
        # 2) Все прямоугольники
        for item in self._boxes:
            self._draw_box_outline(item)
        # 3) Весь текст
        for item in self._boxes:
            self._draw_box_text(item)
        for lbl in self._labels:
            fn = self.font_bold if lbl.get("bold") else self.font_name
            fs = self._label_font if lbl.get("bold") else self._name_font
            fs = fs * float(lbl.get("font_scale") or 1.0)
            canv.setFont(fn, fs)
            lx = lbl["x"] * self.scale_x
            ly = self._ly(lbl["y"])
            if lbl.get("align") == "center":
                canv.drawCentredString(lx, ly, lbl["text"])
            else:
                canv.drawString(lx, ly, lbl["text"])



def _build_bracket_diagram(cat, avail_width, avail_height):
    """Все сетки кумитэ (3 / 4 / 5+ / …) — один чертёж с фиксированными боксами."""
    participants = cat["participants"]
    data = _compute_kumite_bracket_data(participants, cat.get("bouts") or [])
    if not data.get("drawn"):
        return None
    if not data.get("round_robin") and not data.get("rounds_per_group"):
        return None
    level = cat.get("competition_level") or "club"
    return _OfficialBracketDiagram(data, avail_width, avail_height, level)


def _measure_flow_height(flowables, avail_width):
    """Суммарная высота flowables при заданной ширине."""
    total = 0.0
    for f in flowables:
        if isinstance(f, KeepTogether):
            total += _measure_flow_height(f._content, avail_width)
            continue
        _w, h = f.wrap(avail_width, 20000)
        total += h
    return total


def _flatten_flowables(flowables):
    out = []
    for f in flowables:
        if isinstance(f, KeepTogether):
            out.extend(_flatten_flowables(f._content))
        else:
            out.append(f)
    return out


def _build_kumite_protocol_page(styles, tournament, cat, avail_width, usable_height):
    """Шапка + сетка + результаты + подписи — список flowables на один лист."""
    n_part = len(cat.get("participants") or [])
    compact = n_part >= 6

    def make_fixed():
        header = list(_bracket_protocol_header(styles, tournament, cat, compact=compact))
        results = _flatten_flowables([
            _official_bracket_results_table(cat, avail_width, compact=compact)
        ])
        signatures = _bracket_signature_block(tournament, avail_width, compact=compact)
        gap1 = Spacer(1, 0.1 * cm if compact else 0.18 * cm)
        gap2 = Spacer(1, 0.12 * cm if compact else 0.22 * cm)
        return header, gap1, results, gap2, signatures

    header, gap1, results, gap2, signatures = make_fixed()
    fixed_h = _measure_flow_height(header + [gap1] + results + [gap2, signatures], avail_width)
    # Запас: Platypus иногда рисует таблицы чуть выше wrap()
    target = usable_height - 14
    bracket_h = max(target - fixed_h, 2.5 * cm)

    diagram = _build_bracket_diagram(cat, avail_width, bracket_h)
    body = None
    for _ in range(14):
        header, gap1, results, gap2, signatures = make_fixed()
        body = header + ([diagram] if diagram else [
            Paragraph("Жеребьёвка ещё не проведена.", styles["Normal"])
        ]) + [gap1] + results + [gap2, signatures]
        total = _measure_flow_height(body, avail_width)
        if total <= target or not diagram:
            return body
        overflow = total - target
        bracket_h = max(bracket_h - overflow - 10, 2.0 * cm)
        diagram = _build_bracket_diagram(cat, avail_width, bracket_h)

    return body


def build_pdf(tournament, summary, categories, team_ranking):
    """Two of the four sections (итоговый протокол / протокол хода
    соревнования) are laid out per-category to match the federation's own
    protocol forms (see docs/samples); сводная справка and командный зачёт
    stay as simple summary tables since the samples don't cover those."""
    styles = _pdf_styles()
    page_size = landscape(A4)
    margin = 1.2 * cm
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=page_size,
        topMargin=margin, bottomMargin=margin, leftMargin=margin, rightMargin=margin,
    )
    avail_width = page_size[0] - 2 * margin
    usable_height = page_size[1] - 2 * margin

    story = [Paragraph("СпортДок — Итоговые документы", styles["Title"]), Spacer(1, 0.3 * cm)]

    story.append(Paragraph("Сводная справка", styles["Heading2"]))
    info_rows = [
        ["Турнир", tournament.get("name") or ""],
        ["Место проведения", tournament.get("location") or ""],
        ["Дата турнира", tournament.get("event_date") or ""],
        ["Закрытие заявок", tournament.get("registration_closes_at") or ""],
        ["Статус", tournament.get("status") or ""],
        ["Участников заявлено", str(summary.get("participant_count", 0))],
        ["Категорий", str(summary.get("category_count", 0))],
    ]
    story.append(_make_table(info_rows))
    story.append(Spacer(1, 0.3 * cm))
    discipline_rows = [["Дисциплина", "Участников"]] + [
        [DISCIPLINE_LABELS.get(d, d), str(c)] for d, c in summary.get("discipline_counts", {}).items()
    ]
    story.append(_make_table(discipline_rows, header=True))

    for cat in categories:
        # Одна и та же строка категории, что в шапке протокола хода
        program_label = _bracket_program_desc(cat)

        story.append(PageBreak())
        story += _official_header(styles, tournament, program_label, "ИТОГОВЫЙ ПРОТОКОЛ")
        if cat["participants"]:
            story.append(_full_protocol_table(cat, tournament))
        else:
            story.append(Paragraph("Участники не заявлены.", styles["Normal"]))
        story.append(Spacer(1, 1 * cm))
        story.append(_signature_block(styles, tournament))

        story.append(PageBreak())
        if cat["discipline"] == "kata":
            story += _official_header(styles, tournament, program_label, "ПРОТОКОЛ ХОДА СОРЕВНОВАНИЙ")
            kata_table = _kata_rounds_table(cat)
            if kata_table:
                story.append(kata_table)
            else:
                story.append(Paragraph("Результаты пока не введены.", styles["Normal"]))
            story.append(Spacer(1, 0.6 * cm))
            story.append(_signature_block(styles, tournament))
        else:
            story.extend(_build_kumite_protocol_page(
                styles, tournament, cat, avail_width, usable_height))

    story.append(PageBreak())
    story.append(Paragraph("Командный зачёт", styles["Heading2"]))
    if team_ranking:
        rows = [["Клуб", "Очки"]] + [[t["club_name"], str(t["points"])] for t in team_ranking]
        story.append(_make_table(rows, header=True))
    else:
        story.append(Paragraph("Данных пока нет.", styles["Normal"]))

    doc.build(story)
    buffer.seek(0)
    return buffer
