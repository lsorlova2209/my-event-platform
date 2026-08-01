# -*- coding: utf-8 -*-
"""Шаблон заявки по образцу ПР + выпадающие списки кумитэ/ката/пол/разряд."""
from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SRC = Path(r"C:\Users\Liza\Downloads\Заявка ПР 14-20 СПБ (1).xlsx")
OUT_DIR = Path(__file__).resolve().parents[2] / "docs" / "templates"
DESKTOP = Path.home() / "Desktop"
OUT_NAME = "Шаблон_заявки_СпортДок_по_образцу_ПР.xlsx"

DATA_START = 8
DATA_END = 79
EMPTY_SLOTS = DATA_END - DATA_START + 1


def uniq(items: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for x in items:
        x = x.strip()
        if x and x not in seen:
            seen.add(x)
            out.append(x)
    return out


def clear_row_values(ws, row: int, cols: range) -> None:
    for col in cols:
        ws.cell(row, col).value = None


def collect_lists(ref_ws) -> tuple[list[str], list[str], list[str]]:
    kumite: list[str] = []
    kata: list[str] = []
    ranks: list[str] = []
    for r in range(2, (ref_ws.max_row or 1) + 1):
        for col in (6, 7):  # кумитэ 1/2 день
            v = ref_ws.cell(r, col).value
            if v:
                kumite.append(str(v))
        for col in (8, 9, 10):  # ката 1/2/3
            v = ref_ws.cell(r, col).value
            if v:
                kata.append(str(v))
        v = ref_ws.cell(r, 4).value  # разряд
        if v:
            ranks.append(str(v))
    return uniq(kumite), uniq(kata), uniq(ranks)


def write_list_column(ws, col: int, header: str, values: list[str]) -> str:
    """Пишет список в колонку, возвращает формулу диапазона для DataValidation."""
    ws.cell(1, col).value = header
    for i, val in enumerate(values, start=2):
        ws.cell(i, col).value = val
    # очистить хвост старых значений ниже списка
    for r in range(len(values) + 2, max(ws.max_row or 1, len(values) + 2) + 5):
        if ws.cell(r, col).value is not None and r > len(values) + 1:
            # не трогаем чужие данные в старых колонках справочника — только наши L/M/N
            pass
    end_row = 1 + len(values)
    letter = get_column_letter(col)
    return f"Лист1!${letter}$2:${letter}${end_row}"


def main() -> None:
    if not SRC.is_file():
        raise SystemExit(f"Не найден образец: {SRC}")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUT_DIR / OUT_NAME
    shutil.copy2(SRC, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb["Регистрация"]
    ref = wb["Лист1"]

    kumite, kata, ranks = collect_lists(ref)

    # Отдельные колонки справочника для DV (чтобы не упираться в лимит 255 символов)
    # L=12 кумитэ_список, M=13 ката_список, N=14 разряд_список
    kumite_ref = write_list_column(ref, 12, "кумитэ_список", kumite)
    kata_ref = write_list_column(ref, 13, "ката_список", kata)
    ranks_ref = write_list_column(ref, 14, "разряд_список", ranks)

    # Шапка — плейсхолдеры
    ws["B1"] = "ЗАЯВКА на участие"
    ws["B2"] = "в ________________________________ (название турнира)"
    ws["B3"] = "вид спорта: ВСЕСТИЛЕВОЕ КАРАТЭ (0900001411Я)"
    ws["B4"] = "команда РО ФВКР:"
    ws["H4"] = "______________________________"
    ws["B5"] = "место проведения: ____________________"
    ws["J5"] = "дата комиссии по допуску"
    ws["R5"] = None

    for row in range(DATA_START, DATA_END + 1):
        clear_row_values(ws, row, range(2, 20))

    for i in range(EMPTY_SLOTS):
        ws.cell(DATA_START + i, 2).value = i + 1

    # Подписи — убрать конкретные ФИО
    ws["G83"] = "____________________"
    ws["O92"] = "____________________"
    if ws["P86"].value:
        ws["P86"] = "____________________"
    for coord in ("O80", "R80"):
        if ws[coord].value:
            ws[coord] = "/"

    # Сбросить старые валидации и повесить новые на весь диапазон строк
    ws.data_validations.dataValidation = []

    row_range = f"{DATA_START}:{DATA_END}"

    dv_gender = DataValidation(
        type="list",
        formula1='"м,ж"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Пол",
        error="Выберите м или ж",
    )
    dv_gender.add(f"F{DATA_START}:F{DATA_END}")

    dv_kumite = DataValidation(
        type="list",
        formula1=f"={kumite_ref}",
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Кумитэ",
        error="Выберите значение из списка (поединки)",
    )
    # поединки: J K L M
    dv_kumite.add(f"J{DATA_START}:M{DATA_END}")

    dv_kata = DataValidation(
        type="list",
        formula1=f"={kata_ref}",
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Ката",
        error="Выберите значение из списка (ката)",
    )
    # ката: N O P
    dv_kata.add(f"N{DATA_START}:P{DATA_END}")

    dv_rank = DataValidation(
        type="list",
        formula1=f"={ranks_ref}",
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Квалификация",
        error="Выберите разряд из списка",
    )
    dv_rank.add(f"Q{DATA_START}:Q{DATA_END}")

    ws.add_data_validation(dv_gender)
    ws.add_data_validation(dv_kumite)
    ws.add_data_validation(dv_kata)
    ws.add_data_validation(dv_rank)

    wb.save(out_path)

    if DESKTOP.is_dir():
        desktop_path = DESKTOP / OUT_NAME
        try:
            shutil.copy2(out_path, desktop_path)
            print("DESKTOP", desktop_path)
        except PermissionError:
            alt = DESKTOP / OUT_NAME.replace(".xlsx", "_со_списками.xlsx")
            shutil.copy2(out_path, alt)
            print("DESKTOP (alt, файл был открыт):", alt)

    print("OK", out_path)
    print(f"кумитэ: {len(kumite)} значений → {kumite_ref}")
    print(f"ката: {len(kata)} значений → {kata_ref}")
    print(f"разряды: {len(ranks)} значений → {ranks_ref}")
    print(f"диапазон строк {row_range}: пол F, кумитэ J-M, ката N-P, квалиф. Q")


if __name__ == "__main__":
    main()
