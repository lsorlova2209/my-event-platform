# -*- coding: utf-8 -*-
"""Переставить колонки: №, пол, Фамилия, Имя, Отчество."""
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).resolve().parents[2] / "docs" / "templates" / "Шаблон_заявки_СпортДок_по_образцу_ПР.xlsx"
DATA_START, DATA_END = 8, 257


def uniq(items):
    seen, out = set(), []
    for x in items:
        x = str(x).strip()
        if x and x not in seen:
            seen.add(x)
            out.append(x)
    return out


def write_list(ref, col, header, values):
    ref.cell(1, col).value = header
    for i, val in enumerate(values, start=2):
        ref.cell(i, col).value = val
    letter = get_column_letter(col)
    return f"Лист1!${letter}$2:${letter}${1 + len(values)}"


def main():
    wb = openpyxl.load_workbook(OUT)
    ws = wb["Регистрация"]
    ref = wb["Лист1"]

    headers = {
        1: "№",
        2: "пол",
        3: "Фамилия",
        4: "Имя",
        5: "Отчество",
        6: "дата\nрождения",
        7: "полных лет",
        8: "вес",
        9: "Виды программы",
        16: "спорт.\nквалиф.",
        17: "Тренер",
        18: "Допуск\nврача",
    }
    for col, text in headers.items():
        ws.cell(6, col).value = text
        ws.cell(6, col).alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    kumite, kata, ranks = [], [], []
    for r in range(2, (ref.max_row or 1) + 1):
        for col in (6, 7):
            v = ref.cell(r, col).value
            if v:
                kumite.append(v)
        for col in (8, 9, 10):
            v = ref.cell(r, col).value
            if v:
                kata.append(v)
        v = ref.cell(r, 4).value
        if v:
            ranks.append(v)
    kumite, kata, ranks = uniq(kumite), uniq(kata), uniq(ranks)

    kumite_ref = write_list(ref, 12, "кумитэ_список", kumite)
    kata_ref = write_list(ref, 13, "ката_список", kata)
    ranks_ref = write_list(ref, 14, "разряд_список", ranks)

    ws.data_validations.dataValidation = []
    dv_gender = DataValidation(
        type="list", formula1='"м,ж"', allow_blank=True, showDropDown=False,
        showErrorMessage=True, errorTitle="Пол", error="Выберите м или ж",
    )
    dv_gender.add(f"B{DATA_START}:B{DATA_END}")
    dv_kumite = DataValidation(
        type="list", formula1=f"={kumite_ref}", allow_blank=True, showDropDown=False,
        showErrorMessage=True, errorTitle="Кумитэ", error="Выберите из списка",
    )
    dv_kumite.add(f"I{DATA_START}:L{DATA_END}")
    dv_kata = DataValidation(
        type="list", formula1=f"={kata_ref}", allow_blank=True, showDropDown=False,
        showErrorMessage=True, errorTitle="Ката", error="Выберите из списка",
    )
    dv_kata.add(f"M{DATA_START}:O{DATA_END}")
    dv_rank = DataValidation(
        type="list", formula1=f"={ranks_ref}", allow_blank=True, showDropDown=False,
        showErrorMessage=True, errorTitle="Квалификация", error="Выберите разряд",
    )
    dv_rank.add(f"P{DATA_START}:P{DATA_END}")
    for dv in (dv_gender, dv_kumite, dv_kata, dv_rank):
        ws.add_data_validation(dv)

    wb.save(OUT)
    print("OK", [ws.cell(6, c).value for c in range(1, 6)])


if __name__ == "__main__":
    main()
