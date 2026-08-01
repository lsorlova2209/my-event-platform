# -*- coding: utf-8 -*-
"""Привести шаблон заявки ПР к требованиям СпортДок.

- всё со столбца A;
- 250 строк участников;
- без заливок, жёлтым только ячейки названия команды;
- плейсхолдеры названия / места / даты комиссии.
"""
from __future__ import annotations

from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "docs" / "templates" / "Шаблон_заявки_СпортДок_по_образцу_ПР.xlsx"

DATA_START = 8
MAX_ATHLETES = 250
DATA_END = DATA_START + MAX_ATHLETES - 1  # 257

YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
NO_FILL = PatternFill(fill_type=None)


def _unmerge_all(ws) -> None:
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))


def _shift_values_left(ws, max_row: int, max_col: int = 30) -> None:
    """Сдвиг значений B→A … без delete_cols (он ломает merges)."""
    for row in range(1, max_row + 1):
        for col in range(1, max_col):
            src = ws.cell(row, col + 1)
            dst = ws.cell(row, col)
            dst.value = src.value
            if src.has_style:
                dst.font = copy(src.font)
                dst.border = copy(src.border)
                dst.fill = copy(src.fill)
                dst.alignment = copy(src.alignment)
                dst.number_format = src.number_format
        ws.cell(row, max_col).value = None


def _copy_row_style(ws, src_row: int, dst_row: int, max_col: int = 18) -> None:
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
            dst.fill = NO_FILL


def _find_footer_start(ws) -> int:
    for r in range(70, min((ws.max_row or 80) + 1, 200)):
        for c in range(1, 8):
            text = ws.cell(r, c).value
            if not isinstance(text, str):
                continue
            t = text.strip()
            if t.startswith("Представленные документы") or t.startswith("Представленные"):
                return r
            if t.startswith("Тренер команды"):
                return r
    # запасной вариант: первая строка после данных, где в A не номер слота
    for r in range(70, min((ws.max_row or 80) + 1, 200)):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip() and not v.strip().isdigit():
            if "медицинск" in v.lower() or "Подтверждаю" in v or "Достоверность" in v:
                # слишком глубоко в подписях — ищем выше блок «Представленные»
                continue
            return r
    return 80


def _ensure_athlete_rows(ws) -> None:
    footer_start = _find_footer_start(ws)
    current_slots = footer_start - DATA_START
    need = MAX_ATHLETES - current_slots
    if need > 0:
        ws.insert_rows(footer_start, need)
        for i in range(need):
            row = footer_start + i
            _copy_row_style(ws, DATA_START, row)
            for col in range(1, 19):
                ws.cell(row, col).value = None
            ws.cell(row, 1).value = current_slots + i + 1

    for i in range(MAX_ATHLETES):
        row = DATA_START + i
        ws.cell(row, 1).value = i + 1
        for col in range(2, 19):
            ws.cell(row, col).value = None
            if ws.cell(row, col).fill and ws.cell(row, col).fill.fill_type:
                ws.cell(row, col).fill = NO_FILL


def _clear_fills(ws, max_row: int) -> None:
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=20):
        for cell in row:
            if cell.fill and cell.fill.fill_type:
                cell.fill = NO_FILL


def _remerge_header(ws) -> None:
    merges = [
        "A1:R1",
        "A2:R2",
        "A3:R3",
        "A4:F4",
        "G4:O4",  # название команды — жёлтое
        "A5:F5",
        "I5:P5",  # дата комиссии
        "A6:A7",
        "B6:B7",
        "C6:C7",
        "D6:D7",
        "E6:E7",
        "F6:F7",
        "G6:G7",
        "H6:H7",
        "I6:O6",  # виды программы
        "I7:L7",  # поединки
        "M7:O7",  # ката
        "P6:P7",
        "Q6:Q7",
        "R6:R7",
    ]
    for rng in merges:
        try:
            ws.merge_cells(rng)
        except ValueError:
            pass


def _set_headers(ws) -> None:
    ws["A1"] = "ЗАЯВКА на участие"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws["A2"] = "в {{TOURNAMENT_NAME}}"
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws["A3"] = "вид спорта: ВСЕСТИЛЕВОЕ КАРАТЭ (0900001411Я)"
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

    ws["A4"] = "команда РО ФВКР:"
    ws["G4"] = ""
    for col in range(7, 16):  # G..O
        ws.cell(4, col).fill = YELLOW

    ws["A5"] = "место проведения: {{LOCATION}}"
    ws["I5"] = "дата комиссии по допуску: {{ADMISSION_DATE}}"

    headers = {
        1: "№",
        2: "Фамилия",
        3: "Имя",
        4: "Отчество",
        5: "пол",
        6: "дата\nрождения",
        7: "полных лет",
        8: "вес",
        9: "Виды программы",
        16: "спорт.\nквалиф.",
        17: "Тренер",
        18: "Допуск\nврача",
    }
    for col, text in headers.items():
        ws.cell(6, col).value = text
        ws.cell(6, col).alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    ws["I7"] = "поединки"
    ws["M7"] = "ката"
    ws["I7"].alignment = Alignment(horizontal="center")
    ws["M7"].alignment = Alignment(horizontal="center")


def _reset_validations(wb) -> None:
    ws = wb["Регистрация"]
    ref = wb["Лист1"]

    def uniq(items):
        seen, out = set(), []
        for x in items:
            x = str(x).strip()
            if x and x not in seen:
                seen.add(x)
                out.append(x)
        return out

    kumite, kata, ranks = [], [], []
    for r in range(2, (ref.max_row or 1) + 1):
        for col in (6, 7):
            v = ref.cell(r, col).value
            if v:
                kumite.append(v)
        for col in (8, 9, 10):
            v = ref.cell(r, col).value
            if v:
                kata.append(v)
        v = ref.cell(r, 4).value
        if v:
            ranks.append(v)
    kumite, kata, ranks = uniq(kumite), uniq(kata), uniq(ranks)

    def write_list(col, header, values):
        ref.cell(1, col).value = header
        for i, val in enumerate(values, start=2):
            ref.cell(i, col).value = val
        letter = get_column_letter(col)
        return f"Лист1!${letter}$2:${letter}${1 + len(values)}"

    kumite_ref = write_list(12, "кумитэ_список", kumite)
    kata_ref = write_list(13, "ката_список", kata)
    ranks_ref = write_list(14, "разряд_список", ranks)

    ws.data_validations.dataValidation = []

    dv_gender = DataValidation(
        type="list", formula1='"м,ж"', allow_blank=True, showDropDown=False,
        showErrorMessage=True, errorTitle="Пол", error="Выберите м или ж",
    )
    dv_gender.add(f"E{DATA_START}:E{DATA_END}")

    dv_kumite = DataValidation(
        type="list", formula1=f"={kumite_ref}", allow_blank=True, showDropDown=False,
        showErrorMessage=True, errorTitle="Кумитэ", error="Выберите из списка",
    )
    dv_kumite.add(f"I{DATA_START}:L{DATA_END}")

    dv_kata = DataValidation(
        type="list", formula1=f"={kata_ref}", allow_blank=True, showDropDown=False,
        showErrorMessage=True, errorTitle="Ката", error="Выберите из списка",
    )
    dv_kata.add(f"M{DATA_START}:O{DATA_END}")

    dv_rank = DataValidation(
        type="list", formula1=f"={ranks_ref}", allow_blank=True, showDropDown=False,
        showErrorMessage=True, errorTitle="Квалификация", error="Выберите разряд",
    )
    dv_rank.add(f"P{DATA_START}:P{DATA_END}")

    for dv in (dv_gender, dv_kumite, dv_kata, dv_rank):
        ws.add_data_validation(dv)


def _scrub_footer_names(ws) -> None:
    markers = ("Иванов", "Петров", "Сидоров", "Смирнов")
    for r in range(DATA_END + 1, min((ws.max_row or DATA_END) + 1, DATA_END + 50)):
        for c in range(1, 20):
            v = ws.cell(r, c).value
            if isinstance(v, str) and any(m in v for m in markers):
                ws.cell(r, c).value = "____________________"


def main() -> None:
    if not OUT.is_file():
        raise SystemExit(f"Нет файла: {OUT}")

    wb = openpyxl.load_workbook(OUT)
    ws = wb["Регистрация"]

    # Сначала сдвигаем, пока ещё есть подписи на старых местах
    scan_rows = max(ws.max_row or 100, 120)
    _unmerge_all(ws)
    # Колонка A пустая — сдвигаем B→A
    if ws["A1"].value is None and ws["B1"].value is not None:
        _shift_values_left(ws, max_row=scan_rows, max_col=25)

    _ensure_athlete_rows(ws)
    _clear_fills(ws, max_row=max(ws.max_row or DATA_END, DATA_END + 40))
    _set_headers(ws)
    _remerge_header(ws)
    # жёлтая заливка названия команды (после merge — на верхнюю левую и диапазон)
    for col in range(7, 16):
        ws.cell(4, col).fill = YELLOW
    _reset_validations(wb)
    _scrub_footer_names(ws)

    wb.save(OUT)
    print("OK", OUT)
    print(f"athletes {DATA_START}..{DATA_END}")
    print("A1", ws["A1"].value)
    print("A2", ws["A2"].value)
    print("A5", ws["A5"].value)
    print("I5", ws["I5"].value)
    print("G4 yellow", ws["G4"].fill.fgColor.rgb if ws["G4"].fill.fgColor else None)


if __name__ == "__main__":
    main()
