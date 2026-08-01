# -*- coding: utf-8 -*-
"""Заполнить шаблон заявки 72 участниками с квотой ≥10 на категорию в каждой возрастной группе."""
from __future__ import annotations

import random
import shutil
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path

import openpyxl

SRC = Path(__file__).resolve().parents[2] / "docs" / "templates" / "Шаблон_заявки_СпортДок_по_образцу_ПР.xlsx"
DESKTOP = Path.home() / "Desktop"
OUT_NAME = "Шаблон_заявки_СпортДок_по_образцу_ПР_заполненный.xlsx"

REF = date(2026, 4, 1)
DATA_START = 8
PER_AGE = 24  # 3 группы × 24 = 72

# Возрастные группы первенства 14–20
AGE_GROUPS = [
    ("14-15", [14, 15]),
    ("16-17", [16, 17]),
    ("18-20", [18, 19, 20]),
]

# Выбранные весовые (из справочника). 8×10=80 ≤ 24×4=96 слотов на группу.
KUMITE_CATS = [
    "ОК-50", "ОК-55", "ОК-60", "ОК-65", "ОК-70", "ОК-АБС",
    "ПК-60", "СЗ-64",
]

# Все категории ката из справочника. 6×10=60 ≤ 24×3=72 слотов на группу.
KATA_CATS = [
    "ОК-ката-ренгокай",
    "ОК-ката-вадо-рю",
    "ОК-ката-годзю-рю",
    "ОК-ката-группа",
    "СЗ-ката-соло",
    "СЗ-ката-соло с предметом",
]

RANKS_BY_AGE = {
    "14-15": ["3 юн.", "2 юн.", "1 юн.", "3 спорт.", "2 спорт."],
    "16-17": ["1 юн.", "3 спорт.", "2 спорт.", "1 спорт.", "КМС"],
    "18-20": ["2 спорт.", "1 спорт.", "КМС", "МС", "КМС"],
}

TRAINERS = [
    "Орлова Е.С.", "Николаев А.В.", "Воробьева М.А.", "Жихарев Д.М.",
    "Дроздов В.И.", "Аветисов С.О.", "Волков П.Н.", "Суховей В.П.",
    "Любимов О.И.", "Лавренов С.С.", "Пантелеев А.А.", "Омаров Ф.Ш.",
    "Мосенков Е.А.", "Сенцов И.Ю.", "Цыльев В.А.", "Атакишиев И.И.",
]

LAST_M = [
    "Иванов", "Петров", "Сидоров", "Смирнов", "Кузнецов", "Попов", "Васильев",
    "Соколов", "Михайлов", "Новиков", "Фёдоров", "Морозов", "Волков", "Алексеев",
    "Лебедев", "Семёнов", "Егоров", "Павлов", "Козлов", "Степанов", "Николаев",
    "Орлов", "Андреев", "Макаров", "Никитин", "Захаров", "Зайцев", "Соловьёв",
]
LAST_F = [
    "Иванова", "Петрова", "Сидорова", "Смирнова", "Кузнецова", "Попова", "Васильева",
    "Соколова", "Михайлова", "Новикова", "Фёдорова", "Морозова", "Волкова", "Алексеева",
    "Лебедева", "Семёнова", "Егорова", "Павлова", "Козлова", "Степанова", "Николаева",
    "Орлова", "Андреева", "Макарова", "Никитина", "Захарова", "Зайцева", "Соловьёва",
]
FIRST_M = [
    "Александр", "Дмитрий", "Максим", "Иван", "Артём", "Михаил", "Кирилл", "Никита",
    "Андрей", "Егор", "Илья", "Алексей", "Роман", "Сергей", "Владимир", "Павел",
    "Константин", "Тимур", "Матвей", "Даниил", "Глеб", "Лев", "Марк", "Ярослав",
]
FIRST_F = [
    "Анна", "Мария", "Елена", "Дарья", "Полина", "Алина", "Виктория", "София",
    "Ксения", "Анастасия", "Екатерина", "Валерия", "Диана", "Арина", "Милана",
    "Вероника", "Ульяна", "Кира", "Василиса", "Ева", "Злата", "Варвара", "Таисия", "Ольга",
]
MIDDLE_M = [
    "Александрович", "Дмитриевич", "Сергеевич", "Иванович", "Андреевич", "Михайлович",
    "Алексеевич", "Владимирович", "Николаевич", "Павлович", "Романович", "Игоревич",
]
MIDDLE_F = [
    "Александровна", "Дмитриевна", "Сергеевна", "Ивановна", "Андреевна", "Михайловна",
    "Алексеевна", "Владимировна", "Николаевна", "Павловна", "Романовна", "Игоревна",
]

WEIGHT_HINT = {
    "ОК-50": 49.0, "ОК-55": 54.0, "ОК-60": 59.0, "ОК-65": 64.0, "ОК-70": 69.0,
    "ОК-АБС": 72.0, "ПК-60": 59.5, "СЗ-64": 63.0,
}


def birth_for_age(age: int, rng: random.Random) -> date:
    # возраст на REF: age полных лет
    latest = REF.replace(year=REF.year - age)
    earliest = REF.replace(year=REF.year - age - 1) + timedelta(days=1)
    span = (latest - earliest).days
    return earliest + timedelta(days=rng.randint(0, max(span, 0)))


def assign_categories(n_athletes: int, categories: list[str], per_cat: int, max_slots: int, rng: random.Random) -> list[list[str]]:
    """Каждому атлету — список категорий; каждая категория ≥ per_cat раз."""
    bags: list[list[str]] = [[] for _ in range(n_athletes)]
    queue: list[str] = []
    for cat in categories:
        queue.extend([cat] * per_cat)
    rng.shuffle(queue)

    for cat in queue:
        # атлеты без этой категории, с свободным слотом, меньше всего категорий
        candidates = [
            i for i in range(n_athletes)
            if cat not in bags[i] and len(bags[i]) < max_slots
        ]
        if not candidates:
            # снять избыточную категорию у кого-то
            freed = False
            for i in range(n_athletes):
                if cat in bags[i] or len(bags[i]) == 0:
                    continue
                # ищем категорию с избытком
                counts = defaultdict(int)
                for b in bags:
                    for c in b:
                        counts[c] += 1
                removable = [c for c in bags[i] if counts[c] > per_cat]
                if removable:
                    bags[i].remove(removable[0])
                    bags[i].append(cat)
                    freed = True
                    break
            if not freed:
                raise RuntimeError(f"Нет места для {cat}")
            continue
        candidates.sort(key=lambda i: (len(bags[i]), rng.random()))
        bags[candidates[0]].append(cat)

    return bags


def verify(bags: list[list[str]], categories: list[str], per_cat: int) -> dict[str, int]:
    counts = {c: 0 for c in categories}
    for bag in bags:
        for c in bag:
            if c in counts:
                counts[c] += 1
    bad = {c: n for c, n in counts.items() if n < per_cat}
    if bad:
        raise RuntimeError(f"Квота не выполнена: {bad}")
    return counts


def main() -> None:
    if not SRC.is_file():
        raise SystemExit(f"Нет шаблона: {SRC}")

    rng = random.Random(20260401)
    wb = openpyxl.load_workbook(SRC)
    ws = wb["Регистрация"]

    ws["B2"] = "в первенстве России по всестилевому каратэ среди юношей и девушек 14-15 лет, юниоров и юниорок 16-17 лет, 18-20 лет (демо-заполнение)"
    ws["H4"] = "города Санкт-Петербурга"
    ws["B5"] = "место проведения: г. Орёл"
    ws["R5"] = date(2026, 4, 1)

    row = DATA_START
    num = 1
    summary = []

    for group_name, ages in AGE_GROUPS:
        kumite_bags = assign_categories(PER_AGE, KUMITE_CATS, 10, 4, rng)
        kata_bags = assign_categories(PER_AGE, KATA_CATS, 10, 3, rng)
        k_counts = verify(kumite_bags, KUMITE_CATS, 10)
        ka_counts = verify(kata_bags, KATA_CATS, 10)
        summary.append((group_name, k_counts, ka_counts))

        for i in range(PER_AGE):
            gender = "ж" if i % 3 == 0 else "м"  # ~1/3 девушек
            age = ages[i % len(ages)]
            bday = birth_for_age(age, rng)
            if gender == "м":
                last = LAST_M[(num + i) % len(LAST_M)]
                first = FIRST_M[(num * 3 + i) % len(FIRST_M)]
                middle = MIDDLE_M[(num + i * 2) % len(MIDDLE_M)]
            else:
                last = LAST_F[(num + i) % len(LAST_F)]
                first = FIRST_F[(num * 3 + i) % len(FIRST_F)]
                middle = MIDDLE_F[(num + i * 2) % len(MIDDLE_F)]

            k_list = kumite_bags[i][:4]
            ka_list = kata_bags[i][:3]
            primary_weight = WEIGHT_HINT.get(k_list[0], 60.0) if k_list else 55.0
            weight = round(primary_weight + rng.uniform(-1.2, 1.2), 1)
            rank = RANKS_BY_AGE[group_name][i % len(RANKS_BY_AGE[group_name])]
            trainer = TRAINERS[(num + i) % len(TRAINERS)]
            if i % 7 == 0:
                trainer = f"{trainer}, {TRAINERS[(num + i + 3) % len(TRAINERS)]}"

            ws.cell(row, 2).value = num
            ws.cell(row, 3).value = last
            ws.cell(row, 4).value = first
            ws.cell(row, 5).value = middle
            ws.cell(row, 6).value = gender
            ws.cell(row, 7).value = bday
            ws.cell(row, 7).number_format = "DD.MM.YYYY"
            ws.cell(row, 8).value = age
            ws.cell(row, 9).value = weight
            # поединки J-M (10-13)
            for col, cat in zip(range(10, 14), k_list):
                ws.cell(row, col).value = cat
            for col in range(10 + len(k_list), 14):
                ws.cell(row, col).value = None
            # ката N-P (14-16)
            for col, cat in zip(range(14, 17), ka_list):
                ws.cell(row, col).value = cat
            for col in range(14 + len(ka_list), 17):
                ws.cell(row, col).value = None
            ws.cell(row, 17).value = rank
            ws.cell(row, 18).value = trainer
            ws.cell(row, 19).value = None

            row += 1
            num += 1

    out_path = SRC.parent / OUT_NAME
    wb.save(out_path)

    desktop_paths = []
    if DESKTOP.is_dir():
        for name in (OUT_NAME, OUT_NAME.replace(".xlsx", "_v2.xlsx")):
            try:
                p = DESKTOP / name
                shutil.copy2(out_path, p)
                desktop_paths.append(p)
                break
            except PermissionError:
                continue

    print("OK", out_path)
    for p in desktop_paths:
        print("DESKTOP", p)
    print("Всего участников:", num - 1)
    for group_name, k_counts, ka_counts in summary:
        print(f"\n=== {group_name} ===")
        print("кумитэ:", dict(k_counts))
        print("ката:", dict(ka_counts))
        print("min кумитэ:", min(k_counts.values()), "min ката:", min(ka_counts.values()))


if __name__ == "__main__":
    main()
