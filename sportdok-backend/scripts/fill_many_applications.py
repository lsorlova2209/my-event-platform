# -*- coding: utf-8 -*-
"""
Генерирует ≥15 заявок из разных клубов, суммарно ~3000 участников.
Возрасты: 10-11 (только ката), 12-13 / 14-15 / 16-17 / 18-20 (ката+кумитэ).
В каждой возрастной группе заявки: ≥10 на каждую используемую категорию ката/кумитэ.
"""
from __future__ import annotations

import random
import re
import shutil
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

TEMPLATE = Path(__file__).resolve().parents[2] / "docs" / "templates" / "Шаблон_заявки_СпортДок_по_образцу_ПР.xlsx"
OUT_DIR = Path(__file__).resolve().parents[2] / "docs" / "templates" / "applications_3000"
DESKTOP_DIR = Path.home() / "Desktop" / "Заявки_демо_3000"

REF = date(2026, 4, 1)
DATA_START = 8
FOOTER_ROW = 80  # «Врачом допущено» в шаблоне
BASE_SLOTS = 72  # строк 8..79

N_APPLICATIONS = 15
PER_APP = 200  # 15 × 200 = 3000
PER_AGE = PER_APP // 5  # 40 на возрастную группу

AGE_GROUPS = [
    ("10-11", [10, 11], True),   # только ката
    ("12-13", [12, 13], False),
    ("14-15", [14, 15], False),
    ("16-17", [16, 17], False),
    ("18-20", [18, 19, 20], False),
]

KATA_CATS = [
    "ОК-ката-ренгокай",
    "ОК-ката-вадо-рю",
    "ОК-ката-годзю-рю",
    "ОК-ката-группа",
    "СЗ-ката-соло",
    "СЗ-ката-соло с предметом",
]

KUMITE_BY_AGE = {
    "12-13": ["ОК-40", "ОК-45", "ОК-50", "ОК-55", "ПК-40", "ПК-50", "СЗ-39", "СЗ-45"],
    "14-15": ["ОК-50", "ОК-55", "ОК-60", "ОК-65", "ОК-70", "ОК-АБС", "ПК-60", "СЗ-64"],
    "16-17": ["ОК-55", "ОК-60", "ОК-65", "ОК-70", "ОК-73", "ОК-АБС", "ПК-70", "СЗ-72"],
    "18-20": ["ОК-60", "ОК-65", "ОК-70", "ОК-75", "ОК-80", "ОК-АБС", "ПК-75", "СЗ-80"],
}

WEIGHT_HINT = {
    "ОК-40": 39, "ОК-45": 44, "ОК-50": 49, "ОК-55": 54, "ОК-60": 59, "ОК-65": 64,
    "ОК-70": 69, "ОК-73": 72, "ОК-75": 74, "ОК-80": 79, "ОК-АБС": 78,
    "ПК-40": 39, "ПК-50": 49, "ПК-60": 59, "ПК-70": 69, "ПК-75": 74,
    "СЗ-39": 38, "СЗ-45": 44, "СЗ-64": 63, "СЗ-72": 71, "СЗ-80": 79,
}

RANKS_BY_AGE = {
    "10-11": ["б/р", "3 юн.", "2 юн.", "1 юн."],
    "12-13": ["б/р", "3 юн.", "2 юн.", "1 юн.", "3 спорт."],
    "14-15": ["3 юн.", "2 юн.", "1 юн.", "3 спорт.", "2 спорт."],
    "16-17": ["1 юн.", "3 спорт.", "2 спорт.", "1 спорт.", "КМС"],
    "18-20": ["2 спорт.", "1 спорт.", "КМС", "МС", "КМС"],
}

TRAINERS = [
    "Орлова Е.С.", "Николаев А.В.", "Воробьева М.А.", "Жихарев Д.М.",
    "Дроздов В.И.", "Аветисов С.О.", "Волков П.Н.", "Суховей В.П.",
    "Любимов О.И.", "Лавренов С.С.", "Пантелеев А.А.", "Омаров Ф.Ш.",
    "Мосенков Е.А.", "Сенцов И.Ю.", "Цыльев В.А.", "Атакишиев И.И.",
]

LAST_M = [
    "Иванов", "Петров", "Сидоров", "Смирнов", "Кузнецов", "Попов", "Васильев", "Соколов",
    "Михайлов", "Новиков", "Фёдоров", "Морозов", "Волков", "Алексеев", "Лебедев", "Семёнов",
    "Егоров", "Павлов", "Козлов", "Степанов", "Николаев", "Орлов", "Андреев", "Макаров",
    "Никитин", "Захаров", "Зайцев", "Соловьёв", "Борисов", "Яковлев", "Григорьев", "Романов",
]
LAST_F = [
    "Иванова", "Петрова", "Сидорова", "Смирнова", "Кузнецова", "Попова", "Васильева", "Соколова",
    "Михайлова", "Новикова", "Фёдорова", "Морозова", "Волкова", "Алексеева", "Лебедева", "Семёнова",
    "Егорова", "Павлова", "Козлова", "Степанова", "Николаева", "Орлова", "Андреева", "Макарова",
    "Никитина", "Захарова", "Зайцева", "Соловьёва", "Борисова", "Яковлева", "Григорьева", "Романова",
]
FIRST_M = [
    "Александр", "Дмитрий", "Максим", "Иван", "Артём", "Михаил", "Кирилл", "Никита",
    "Андрей", "Егор", "Илья", "Алексей", "Роман", "Сергей", "Владимир", "Павел",
    "Константин", "Тимур", "Матвей", "Даниил", "Глеб", "Лев", "Марк", "Ярослав",
    "Тимофей", "Фёдор", "Савелий", "Демид", "Платон", "Гордей", "Мирон", "Эмиль",
]
FIRST_F = [
    "Анна", "Мария", "Елена", "Дарья", "Полина", "Алина", "Виктория", "София",
    "Ксения", "Анастасия", "Екатерина", "Валерия", "Диана", "Арина", "Милана",
    "Вероника", "Ульяна", "Кира", "Василиса", "Ева", "Злата", "Варвара", "Таисия", "Ольга",
    "Александра", "Юлия", "Маргарита", "Алиса", "Есения", "Мирослава", "Стефания", "Лилия",
]
MIDDLE_M = [
    "Александрович", "Дмитриевич", "Сергеевич", "Иванович", "Андреевич", "Михайлович",
    "Алексеевич", "Владимирович", "Николаевич", "Павлович", "Романович", "Игоревич",
]
MIDDLE_F = [
    "Александровна", "Дмитриевна", "Сергеевна", "Ивановна", "Андреевна", "Михайловна",
    "Алексеевна", "Владимировна", "Николаевна", "Павловна", "Романовна", "Игоревна",
]


def uniq_keep(xs):
    s, o = set(), []
    for x in xs:
        if x not in s:
            s.add(x)
            o.append(x)
    return o


def load_clubs(wb) -> list[str]:
    ref = wb["Лист1"]
    clubs = []
    for r in range(2, (ref.max_row or 1) + 1):
        v = ref.cell(r, 1).value
        if v:
            clubs.append(str(v).strip())
    return uniq_keep(clubs)


def birth_for_age(age: int, rng: random.Random) -> date:
    latest = REF.replace(year=REF.year - age)
    earliest = REF.replace(year=REF.year - age - 1) + timedelta(days=1)
    span = (latest - earliest).days
    return earliest + timedelta(days=rng.randint(0, max(span, 0)))


def assign_categories(n: int, categories: list[str], per_cat: int, max_slots: int, rng: random.Random) -> list[list[str]]:
    bags: list[list[str]] = [[] for _ in range(n)]
    queue: list[str] = []
    for cat in categories:
        queue.extend([cat] * per_cat)
    rng.shuffle(queue)

    for cat in queue:
        candidates = [i for i in range(n) if cat not in bags[i] and len(bags[i]) < max_slots]
        if not candidates:
            counts = defaultdict(int)
            for b in bags:
                for c in b:
                    counts[c] += 1
            freed = False
            for i in range(n):
                if cat in bags[i]:
                    continue
                removable = [c for c in bags[i] if counts[c] > per_cat]
                if removable:
                    bags[i].remove(removable[0])
                    bags[i].append(cat)
                    freed = True
                    break
            if not freed:
                raise RuntimeError(f"Нет места для {cat}")
            continue
        candidates.sort(key=lambda i: (len(bags[i]), rng.random()))
        bags[candidates[0]].append(cat)
    return bags


def safe_name(s: str) -> str:
    s = re.sub(r'[\\/:*?"<>|]+', "_", s)
    return s[:80].strip(" ._") or "club"


def ensure_rows(ws, need_slots: int) -> int:
    """Расширяет таблицу до need_slots строк данных. Возвращает DATA_END."""
    extra = need_slots - BASE_SLOTS
    if extra > 0:
        ws.insert_rows(FOOTER_ROW, amount=extra)
    return DATA_START + need_slots - 1


def reapply_validations(ws, data_end: int, ref) -> None:
    # списки на Лист1 в L/M/N
    def last_row(col: int) -> int:
        r = 2
        while ref.cell(r, col).value:
            r += 1
        return r - 1

    k_end = last_row(12)
    ka_end = last_row(13)
    r_end = last_row(14)

    ws.data_validations.dataValidation = []
    dv_g = DataValidation(type="list", formula1='"м,ж"', allow_blank=True, showDropDown=False)
    dv_g.add(f"F{DATA_START}:F{data_end}")
    dv_k = DataValidation(type="list", formula1=f"=Лист1!$L$2:$L${k_end}", allow_blank=True, showDropDown=False)
    dv_k.add(f"J{DATA_START}:M{data_end}")
    dv_ka = DataValidation(type="list", formula1=f"=Лист1!$M$2:$M${ka_end}", allow_blank=True, showDropDown=False)
    dv_ka.add(f"N{DATA_START}:P{data_end}")
    dv_r = DataValidation(type="list", formula1=f"=Лист1!$N$2:$N${r_end}", allow_blank=True, showDropDown=False)
    dv_r.add(f"Q{DATA_START}:Q{data_end}")
    for dv in (dv_g, dv_k, dv_ka, dv_r):
        ws.add_data_validation(dv)


def fill_application(template_path: Path, club: str, seed: int, out_path: Path) -> int:
    rng = random.Random(seed)
    wb = openpyxl.load_workbook(template_path)
    ws = wb["Регистрация"]

    data_end = ensure_rows(ws, PER_APP)
    reapply_validations(ws, data_end, wb["Лист1"])

    ws["B1"] = "ЗАЯВКА на участие"
    ws["B2"] = "в первенстве / всероссийских соревнованиях по всестилевому каратэ (демо-набор 3000)"
    ws["B3"] = "вид спорта: ВСЕСТИЛЕВОЕ КАРАТЭ (0900001411Я)"
    ws["B4"] = "команда РО ФВКР:"
    ws["H4"] = club
    ws["B5"] = "место проведения: г. Орёл"
    ws["J5"] = "дата комиссии по допуску"
    ws["R5"] = date(2026, 4, 1)

    # очистить старые демо-строки
    for r in range(DATA_START, data_end + 1):
        for c in range(2, 20):
            ws.cell(r, c).value = None

    row = DATA_START
    num = 1
    for group_name, ages, kata_only in AGE_GROUPS:
        kata_bags = assign_categories(PER_AGE, KATA_CATS, 10, 3, rng)
        kumite_bags = (
            [[] for _ in range(PER_AGE)]
            if kata_only
            else assign_categories(PER_AGE, KUMITE_BY_AGE[group_name], 10, 4, rng)
        )

        for i in range(PER_AGE):
            gender = "ж" if i % 3 == 0 else "м"
            age = ages[i % len(ages)]
            bday = birth_for_age(age, rng)
            if gender == "м":
                last = LAST_M[(seed + num) % len(LAST_M)]
                first = FIRST_M[(seed * 5 + num) % len(FIRST_M)]
                middle = MIDDLE_M[(seed + num * 2) % len(MIDDLE_M)]
            else:
                last = LAST_F[(seed + num) % len(LAST_F)]
                first = FIRST_F[(seed * 5 + num) % len(FIRST_F)]
                middle = MIDDLE_F[(seed + num * 2) % len(MIDDLE_F)]
            # чуть уникальности по клубу
            if num % 11 == 0:
                last = f"{last}-{club[:3]}"

            k_list = kumite_bags[i][:4]
            ka_list = kata_bags[i][:3]
            if k_list:
                base_w = WEIGHT_HINT.get(k_list[0], 50.0)
            else:
                base_w = 35.0 + age  # для ката-only
            weight = round(base_w + rng.uniform(-1.0, 1.0), 1)
            rank = RANKS_BY_AGE[group_name][i % len(RANKS_BY_AGE[group_name])]
            trainer = TRAINERS[(seed + num) % len(TRAINERS)]

            ws.cell(row, 2).value = num
            ws.cell(row, 3).value = last
            ws.cell(row, 4).value = first
            ws.cell(row, 5).value = middle
            ws.cell(row, 6).value = gender
            ws.cell(row, 7).value = bday
            ws.cell(row, 7).number_format = "DD.MM.YYYY"
            ws.cell(row, 8).value = age
            ws.cell(row, 9).value = weight
            for col, cat in zip(range(10, 14), k_list):
                ws.cell(row, col).value = cat
            for col, cat in zip(range(14, 17), ka_list):
                ws.cell(row, col).value = cat
            ws.cell(row, 17).value = rank
            ws.cell(row, 18).value = trainer

            row += 1
            num += 1

    # подписи
    for coord in ("G83", "O92", "P86"):
        # после insert_rows номера сдвинулись на extra
        pass
    extra = PER_APP - BASE_SLOTS
    # старые координаты подписей + extra
    for base, val in (("G83", "____________________"), ("O92", "____________________"), ("P86", "____________________")):
        col = base[0]
        r = int(base[1:]) + max(extra, 0)
        ws[f"{col}{r}"] = val

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return num - 1


def main() -> None:
    if not TEMPLATE.is_file():
        raise SystemExit(f"Нет шаблона: {TEMPLATE}")

    probe = openpyxl.load_workbook(TEMPLATE)
    clubs = load_clubs(probe)[:N_APPLICATIONS]
    probe.close()
    if len(clubs) < N_APPLICATIONS:
        raise SystemExit(f"В справочнике только {len(clubs)} команд, нужно {N_APPLICATIONS}")

    if OUT_DIR.exists():
        shutil.rmtree(OUT_DIR)
    OUT_DIR.mkdir(parents=True)
    if DESKTOP_DIR.exists():
        shutil.rmtree(DESKTOP_DIR)
    DESKTOP_DIR.mkdir(parents=True)

    total = 0
    manifest = []
    for i, club in enumerate(clubs, start=1):
        fname = f"{i:02d}_Заявка_{safe_name(club)}.xlsx"
        out = OUT_DIR / fname
        n = fill_application(TEMPLATE, club, seed=1000 + i * 17, out_path=out)
        shutil.copy2(out, DESKTOP_DIR / fname)
        total += n
        manifest.append(f"{fname}\t{club}\t{n}")
        print(f"[{i}/{N_APPLICATIONS}] {club}: {n}")

    summary = OUT_DIR / "README.txt"
    text = (
        f"Демо-заявки для турнира\n"
        f"Заявок: {N_APPLICATIONS}\n"
        f"Участников суммарно: {total}\n"
        f"На заявку: {PER_APP} (по {PER_AGE} в группах 10-11, 12-13, 14-15, 16-17, 18-20)\n"
        f"10-11 лет — только ката; с 12-13 — ката и кумитэ.\n"
        f"В каждой возрастной группе заявки ≥10 на используемые категории.\n\n"
        + "\n".join(manifest)
        + "\n"
    )
    summary.write_text(text, encoding="utf-8")
    shutil.copy2(summary, DESKTOP_DIR / "README.txt")

    print("TOTAL", total)
    print("DIR", OUT_DIR)
    print("DESKTOP", DESKTOP_DIR)


if __name__ == "__main__":
    main()
